import tkinter
import pickle
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import psycopg2
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime


try:
    conn = psycopg2.connect('dbname=Personnel_Management user=postgres password=2574 host=localhost port=5432')
    print("Успешное подключение к базе данных PostgreSQL")
except Exception as e:
    print(f"Ошибка подключения к базе данных PostgreSQL: {e}")


def department():
    # Создаем графический интерфейс

    root = tk.Tk()
    root.title("Departmens")
    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('ID_отдела', 'Department_name', 'Employees_count')
    tree.heading('#0', text='№')
    tree.heading('ID_отдела', text='ID_department')
    tree.heading('Department_name', text='Name')
    tree.heading('Employees_count', text='Count')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID_отдела", width=100, anchor=tk.CENTER)
    tree.column("Department_name", width=170, anchor=tk.CENTER)
    tree.column("Employees_count", width=50, anchor=tk.CENTER)

    tree.pack()

    try:
        # удаление устаревших данных
        tree.delete(*tree.get_children())

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM department")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"department_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")

    # Добавление данных в таблицу из бд
    def insert():
        try:
            # удаление устаревших данных
            tree.delete(*tree.get_children())

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM department")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    # Добавление новой строки в таблицу
    def add_string():
        name = entry.get()
        location = entry2.get()
        manager = entry3.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO department (\"ID_отдела\", Название_отдела, Количество_сотрудников)"
                           " VALUES (%s, %s, %s)",
                           (name, location, manager))
            conn.commit()

            insert()  # Обновить вывод таблицы после добавления строки
            entry.delete(0, tk.END)
            entry2.delete(0, tk.END)
            entry3.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root.bind('<Return>', on_enter)

    def delete_all_records():
        try:
            selected_item = tree.selection()

            for item in selected_item:
                values = tree.item(item, 'values')
                ID_department = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                try:

                    # Удалить из department_statistics
                    cursor.execute("DELETE FROM department_statistics WHERE \"ID_отдела\" = %s", (ID_department,))

                    # Удалить из employees, ссылка на которые присутствует в department
                    cursor.execute("DELETE FROM employees WHERE \"ID_отдела\" = %s", (ID_department,))

                    # Удаление данных из таблицы 'department'
                    cursor.execute("DELETE FROM department WHERE \"ID_отдела\" = %s", (ID_department,))

                    conn.commit()

                    insert()  # Обновить вывод таблицы после удаления всех записей
                except Exception as e:
                    conn.rollback()
                    error_msg = f"Ошибка при удалении данных из таблиц: Убедитесь, что вы удалили всех сотрудников из данного отдела {e}"
                    print(error_msg)
                    messagebox.showerror("Ошибка", error_msg)
                    return

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при удалении данных из таблиц: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_all_records()
        else:
            print("Action canceled.")

    def update_string():
        id_value = entry.get()
        name = entry2.get()
        count = entry3.get()

        if not id_value or not count:
            messagebox.showerror("Ошибка",
                                 "Пожалуйста, введите корректные значения для ID отдела и количества сотрудников.")
            return

        try:
            id_value = int(id_value)
            count = int(count)

            cursor = conn.cursor()
            cursor.execute(
                "UPDATE department SET Название_отдела = %s, Количество_сотрудников = %s WHERE \"ID_отдела\" = %s",
                (name, count, id_value))
            conn.commit()

            # Обновляем поля после успешного обновления
            entry.delete(0, tk.END)
            entry.insert(0, id_value)
            entry2.delete(0, tk.END)
            entry3.delete(0, tk.END)

            insert()  # Может потребоваться обновить данные на экране
            entry.delete(0, tk.END)

        except ValueError:
            messagebox.showerror("Ошибка", "ID отдела и количество сотрудников должны быть числовыми значениями.")
        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        return entry, entry2, entry3

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле
            entry3.delete(0, tk.END)
            entry3.insert(0, item_values[2])  # Пример: Количество сотрудников в третье поле

    tree.bind("<Double-1>", on_double_click)


    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    def delete1():
        entry.delete(0, tk.END)

    def delete2():
        entry2.delete(0, tk.END)

    def delete3():
        entry3.delete(0, tk.END)

    entry_frame10 = tk.Frame(root, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_string,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame11 = tk.Frame(root, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame12 = tk.Frame(root, background="#D2B4DE")
    entry_frame12.pack(side='right')
    # Создание кнопки "x"
    btn1 = tk.Button(entry_frame12, text="x", command=delete1,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=4, pady=1)
    btn2 = tk.Button(entry_frame12, text="x", command=delete2,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=2, pady=1)
    btn3 = tk.Button(entry_frame12, text="x", command=delete3,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=2, pady=1)

    # Создание строки для ввода новых данных
    entry_frame = tk.Frame(root)
    entry_frame.pack()

    entry_label = tk.Label(entry_frame, text="Введите id отдела:", width=23, height=1, background="#D2B4DE", anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame, width=30)  # Создаем поле ввода шириной 30 символов
    entry.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame2 = tk.Frame(root, background="#D2B4DE")
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите название отдела:", width=23, height=2, background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=30)  # Создаем поле ввода шириной 30 символов
    entry2.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame3 = tk.Frame(root)
    entry_frame3.pack()

    entry_label3 = tk.Label(entry_frame3, text="Введите кол-во сотрудников:", width=23, height=1, background="#D2B4DE",
                            anchor='w')
    entry_label3.pack(side='left')

    entry3 = tk.Entry(entry_frame3, width=30)  # Создаем поле ввода шириной 30 символов
    entry3.pack(side='left')  # Размещаем поле ввода на frame1

    return entry, entry2, entry3  # Возвращаем созданное поле ввода


def accounting():

    # Создаем графический интерфейс
    root1 = tk.Tk()
    root1.title("Accounting")
    root1.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root1.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблиц

    # Создание таблицы Отдел
    tree = ttk.Treeview(root1)
    tree["columns"] = ('ID employees', 'Salary', 'Reward')
    tree.heading('#0', text='№')
    tree.heading('ID employees', text='ID employees')
    tree.heading('Salary', text='Salary')
    tree.heading('Reward', text='Reward')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID employees", width=100, anchor=tk.CENTER)
    tree.column("Salary", width=60, anchor=tk.CENTER)
    tree.column("Reward", width=60, anchor=tk.CENTER)

    tree.pack()

    try:
        tree.delete(*tree.get_children())  # удаление устаревших данных

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM accounting")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")

    def sort_by_accounting():
        try:
            selected_department_id = entry4.get()

            sorted_accounting = []
            for child in tree.get_children():
                values = tree.item(child, 'values')
                if values and values[1] == selected_department_id:
                    sorted_accounting.append(values)

            #Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root1)
            result_window.title("Sorted accounting")
            result_window.resizable(False, False)
            result_window.configure(bg = '#D2B4DE')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"]=('ID employees', 'Salary', 'Reward')
            result_tree.heading('#0', text = '№')
            result_tree.heading('ID employees', text = 'ID employees')
            result_tree.heading('Salary', text='Salary')
            result_tree.heading('Reward', text = 'Reward')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("ID employees", width=100, anchor=tk.CENTER)
            result_tree.column("Salary", width=60, anchor=tk.CENTER)
            result_tree.column("Reward", width=60, anchor=tk.CENTER)

            for i, row in enumerate(sorted_accounting):
                result_tree.insert("", "end", text = str(i), values=row)

            result_tree.pack()

            # Функция сохранения результатов сортировки
            def save_to_excel():
                wb = Workbook()
                ws = wb.active
                for i, column in enumerate(
                        ('ID employees', 'Salary', 'Reward')):
                    ws.cell(row=1, column=i + 1, value=column)

                for idx, account in enumerate(sorted_accounting):
                    for i, value in enumerate(account):
                        ws.cell(row=idx + 2, column=i + 1, value=value)

                desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                save_path = os.path.join(desktop_path, f"accounting_data_{timestamp}.xlsx")

                wb.save(save_path)
                print("Данные успешно сохранены в файле employees_data.xlsx. на рабочий стол")

            # Создание кнопки "Сохранить"
            save_button = tk.Button(result_window, text="Save", command=save_to_excel, bg='#E8DAEF', width=10, height=1)
            save_button.pack()

        except Exception as e:
            error_msg = f"Ошибка при сортировке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"accounting_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")


    def insert2():
        try:
            tree.delete(*tree.get_children())  # удаление устаревших данных

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM accounting")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def add_string():
        name = entry.get()
        location = entry2.get()
        manager = entry3.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO accounting (\"ID_персонала\", Заработная_плата, Премия)"
                           " VALUES (%s, %s, %s)",
                           (name, location, manager))
            conn.commit()

            insert2()  # Обновить вывод таблицы после добавления строки
            entry.delete(0, tk.END)
            entry2.delete(0, tk.END)
            entry3.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root1.bind('<Return>', on_enter)

    def delete_record():
        try:
            selected_item = tree.selection()
            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                Accounting = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

            # Выполнение SQL-запроса для удаления строки по определенному идентификатору
            cursor.execute("DELETE FROM accounting WHERE \"ID_персонала\" = %s", (Accounting,))

            conn.commit()
            insert2()  # Обновить вывод таблицы после добавления строки
            print("Строка успешно удалена из базы данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при удалении строки из базы данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_record()
        else:
            print("Action canceled.")

    def update_record():
        try:
            selected_item = tree.selection()

            if not selected_item:
                messagebox.showinfo("Информация", "Пожалуйста, выберите строку для изменения.")
                return

            name = entry.get()
            location = entry2.get()
            manager = entry3.get()

            if not name or not location or not manager:
                messagebox.showinfo("Информация", "Пожалуйста, заполните все поля для обновления.")
                return

            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                accounting_id = values[0]

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для обновления строки по определенному идентификатору
                cursor.execute(
                    "UPDATE accounting SET \"ID_персонала\" = %s, Заработная_плата = %s, Премия = %s WHERE \"ID_персонала\" = %s",
                    (name, location, manager, accounting_id))

            conn.commit()
            insert2()  # Обновить вывод таблицы после обновления строки
            print("Запись успешно обновлена в базе данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении строки в базе данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        # Очистка полей ввода после обновления
        entry.delete(0, tk.END)
        entry2.delete(0, tk.END)
        entry3.delete(0, tk.END)

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле
            entry3.delete(0, tk.END)
            entry3.insert(0, item_values[2])  # Пример: Количество сотрудников в третье поле

    tree.bind("<Double-1>", on_double_click)

    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    def delete1():
        entry.delete(0, tk.END)

    def delete2():
        entry2.delete(0, tk.END)

    def delete3():
        entry3.delete(0, tk.END)

    def delete4():
        entry4.delete(0, tk.END)

    entry_frame10 = tk.Frame(root1, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_record,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame11 = tk.Frame(root1, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert2,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=10, pady=2)
    #Создание кнопки "сортировать"
    btn7 = tk.Button(entry_frame11, text="Sort", padx=5, pady=1, width=10, height=1, command=sort_by_accounting,
                     bg='#E8DAEF')
    btn7.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame12 = tk.Frame(root1, background="#D2B4DE")
    entry_frame12.pack(side='right')
    # Создание кнопки "x"
    btn1 = tk.Button(entry_frame12, text="x", command=delete1,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=4, pady=1)
    btn2 = tk.Button(entry_frame12, text="x", command=delete2,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=4, pady=1)
    btn3 = tk.Button(entry_frame12, text="x", command=delete3,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=4, pady=1)
    btn4 = tk.Button(entry_frame12, text="x", command=delete4,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=4, pady=1)

    # Создание строк для ввода данных
    entry_frame = tk.Frame(root1)
    entry_frame.pack()

    entry_label = tk.Label(entry_frame, text="Введите id сотрудника:", width=18, height=1, background="#D2B4DE",
                           anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame, width=30)  # Создаем поле ввода шириной 30 символов
    entry.pack(side='left')  # Размещаем поле ввода на frame

    entry_frame2 = tk.Frame(root1, background="#D2B4DE")
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите зарплату:", width=18, height=2, background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=30)  # Создаем поле ввода шириной 30 символов
    entry2.pack(side='left')  # Размещаем поле ввода на frame2

    entry_frame3 = tk.Frame(root1, background="#D2B4DE")
    entry_frame3.pack()

    entry_label3 = tk.Label(entry_frame3, text="Введите премию:", width=18, height=1, background="#D2B4DE", anchor='w')
    entry_label3.pack(side='left')

    entry3 = tk.Entry(entry_frame3, width=30)  # Создаем поле ввода шириной 30 символов
    entry3.pack(side='left')  # Размещаем поле ввода на frame3

    entry_frame4 = tk.Frame(root1, background="#D2B4DE")
    entry_frame4.pack()

    entry_label4 = tk.Label(entry_frame4, text="Введите з/п сорт.:", width=18, height=2, background="#D2B4DE", anchor='w')
    entry_label4.pack(side='left')

    entry4 = tk.Entry(entry_frame4, width=30, bg="light gray")  # Создаем поле ввода шириной 30 символов
    entry4.pack(side='left')  # Размещаем поле ввода на frame3

    return entry,  entry2, entry3, entry4  # Возвращаем созданное поле ввода


def department_statistics():
    # Создаем графический интерфейс
    root2 = tk.Tk()
    root2.title("Department statistics")
    root2.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root2.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблиц

    # Создание таблицы Отдел
    tree = ttk.Treeview(root2)
    tree["columns"] = ('ID department', '% work')
    tree.heading('#0', text='№')
    tree.heading('ID department', text='ID_department')
    tree.heading('% work', text='% work')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID department", width=100, anchor=tk.CENTER)
    tree.column("% work", width=60, anchor=tk.CENTER)

    tree.pack()

    try:
        tree.delete(*tree.get_children())  # удаление устаревших данных

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM department_statistics")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row, )

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")
    # Создание кнопок

    def sort_by_percentage():
        try:
            selected_department_id = entry3.get()

            sorted_employees = []
            for child in tree.get_children():
                values = tree.item(child, 'values')
                if values and values[1] == selected_department_id:
                    sorted_employees.append(values)

            # Создаем новое окно для отображения отсортированных данных
            result_window = tk.Toplevel(root2)
            result_window.title("Sorted Department Statistics")
            result_window.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
            result_window.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = ('ID department', '% work')
            result_tree.heading('#0', text='№')
            result_tree.heading('ID department', text='ID department')
            result_tree.heading('% work', text='% work')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("ID department", width=100, anchor=tk.CENTER)
            result_tree.column("% work", width=60, anchor=tk.CENTER)

            for i, row in enumerate(sorted_employees):
                result_tree.insert("", "end", text=str(i), values=row)

            result_tree.pack()

            # Функция сохранения результатов сортировки
            def save_to_excel():
                wb = Workbook()
                ws = wb.active
                for i, column in enumerate(
                        ('ID employees', 'FIO', 'Job title', 'Number phone', 'Email', 'ID department')):
                    ws.cell(row=1, column=i + 1, value=column)

                for idx, employee in enumerate(sorted_employees):
                    for i, value in enumerate(employee):
                        ws.cell(row=idx + 2, column=i + 1, value=value)

                desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                save_path = os.path.join(desktop_path, f"Statistics_data_{timestamp}.xlsx")

                wb.save(save_path)
                print("Данные успешно сохранены в файле Statistics_data.xlsx. на рабочий стол")

            # Создание кнопки "Сохранить"
            save_button = tk.Button(result_window, text="Save", command=save_to_excel, bg='#E8DAEF', width=10, height=1)
            save_button.pack()

        except Exception as e:
            error_msg = f"Ошибка при сортировке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"department_statistic_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")


    def insert3():
        try:
            tree.delete(*tree.get_children())  # удаление устаревших данных

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM department_statistics")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row,)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def add_string():
        name = entry2.get()
        location = entry2.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO department_statistics (\"ID_отдела\", Процент_выполненной_и_невыполненн)"
                           " VALUES (%s, %s)",
                           (name, location))
            conn.commit()

            insert3()  # Обновить вывод таблицы после добавления строки
            entry2.delete(0, tk.END)
            entry2.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root2.bind('<Return>', on_enter)

    def delete_record():
        try:
            selected_item = tree.selection()
            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                Statistic = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

            # Выполнение SQL-запроса для удаления строки по определенному идентификатору
            cursor.execute("DELETE FROM department_statistics WHERE \"ID_отдела\" = %s", (Statistic,))

            conn.commit()
            insert3()  # Обновить вывод таблицы после добавления строки
            print("Строка успешно удалена из базы данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при удалении строки из базы данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_record()
        else:
            print("Action canceled.")

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле

    tree.bind("<Double-1>", on_double_click)

    def update_record():
        try:
            selected_item = tree.selection()

            if not selected_item:
                messagebox.showinfo("Информация", "Пожалуйста, выберите строку для изменения.")
                return

            name = entry.get()
            percentage_work = entry2.get()

            if not name or not percentage_work:
                messagebox.showinfo("Информация", "Пожалуйста, заполните все поля для обновления.")
                return

            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                department_id = values[0]

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для обновления строки по определенному идентификатору
                cursor.execute(
                    "UPDATE department_statistics SET \"ID_отдела\" = %s, Процент_выполненной_и_невыполненн = %s WHERE \"ID_отдела\" = %s",
                    (name, percentage_work, department_id))

            conn.commit()
            insert3()  # Обновить вывод таблицы после обновления строки
            print("Запись успешно обновлена в базе данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении строки в базе данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        # Очистка полей ввода после обновления
        entry.delete(0, tk.END)
        entry2.delete(0, tk.END)

    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    entry_frame10 = tk.Frame(root2, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_record,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame11 = tk.Frame(root2, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert3,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "СОРТИРОВАТЬ"
    btn6 = tk.Button(entry_frame11, text="Sort", padx=5, pady=1, width=10, height=1, command=sort_by_percentage,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=10, pady=2)

    entry_frame12 = tk.Frame(root2, background="#D2B4DE")
    entry_frame12.pack(side='right')

    entry_frame = tk.Frame(root2, pady=1, background="#D2B4DE")
    entry_frame.pack()

    entry_label = tk.Label(entry_frame, text="Введите id отдела:", width=16, height=2, background="#D2B4DE", anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame, width=30)  # Создаем поле ввода шириной 30 символов
    entry.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame2 = tk.Frame(root2, background="#D2B4DE", pady=1)
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите % работы:", width=16, height=1, background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=30)  # Создаем поле ввода шириной 30 символов
    entry2.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame3 = tk.Frame(root2, background="#D2B4DE")
    entry_frame3.pack()

    entry_label3 = tk.Label(entry_frame3, text="Сортировка по %:", width=16, height=2, background="#D2B4DE",
                            anchor='w')
    entry_label3.pack(side='left')

    entry3 = tk.Entry(entry_frame3, width=30, bg='light gray')  # Создаем поле ввода шириной 30 символов
    entry3.pack(side='left')  # Размещаем поле ввода на frame1

    return entry, entry2, entry3  # Возвращаем созданное поле ввода


def employees():
    # Создаем графический интерфейс
    root3 = tk.Tk()
    root3.title("Employees")
    root3.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root3.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root3)
    tree["columns"] = ('ID employees', 'FIO', 'Job title', 'Number phone', 'Email', 'ID department')
    tree.heading('#0', text='№')
    tree.heading('ID employees', text='ID employees')
    tree.heading('FIO', text='FIO')
    tree.heading('Job title', text='Job title')
    tree.heading('Number phone', text='Number phone')
    tree.heading('Email', text='Email')
    tree.heading('ID department', text='ID department')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID employees", width=100, anchor=tk.CENTER)
    tree.column("FIO", width=200, anchor=tk.CENTER)
    tree.column("Job title", width=150, anchor=tk.CENTER)
    tree.column("Number phone", width=100, anchor=tk.CENTER)
    tree.column("Email", width=150, anchor=tk.CENTER)
    tree.column("ID department", width=100, anchor=tk.CENTER)

    tree.pack()

    try:
        tree.delete(*tree.get_children())  # удаление устаревших данных

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM employees")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row, )

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")

    # Функция для сортировки по отделу и вывода результата в отдельное окно
    def sort_by_department():
        selected_department_id = sort_entry.get()

        sorted_employees = []
        for child in tree.get_children():
            values = tree.item(child, 'values')
            if values and values[5] == selected_department_id:
                sorted_employees.append(values)

        # Создание нового окна для вывода отсортированных данных
        result_window = tk.Toplevel(root3)
        result_window.title("Sorted Employees by Department")
        result_window.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
        result_window.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

        result_tree = ttk.Treeview(result_window)
        result_tree["columns"] = ('ID employees', 'FIO', 'Job title', 'Number phone', 'Email', 'ID department')

        # Создание колонок таблицы
        for i, column in enumerate(('ID employees', 'FIO', 'Job title', 'Number phone', 'Email', 'ID department')):
            result_tree.heading(i, text=column)
            result_tree.column(i, width=180, minwidth=100, anchor=tk.CENTER, stretch=True)

        for idx, employee in enumerate(sorted_employees):
            result_tree.insert("", "end", text=str(idx), values=employee)

        result_tree.pack()

        # Функция сохранения результатов сортировки
        def save_to_excel():
            wb = Workbook()
            ws = wb.active
            for i, column in enumerate(('ID employees', 'FIO', 'Job title', 'Number phone', 'Email', 'ID department')):
                ws.cell(row=1, column=i + 1, value=column)

            for idx, employee in enumerate(sorted_employees):
                for i, value in enumerate(employee):
                    ws.cell(row=idx + 2, column=i + 1, value=value)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"employees_data_{timestamp}.xlsx")

            wb.save(save_path)
            print("Данные успешно сохранены в файле employees_data.xlsx. на рабочий стол")

        # Создание кнопки "Сохранить"
        save_button = tk.Button(result_window, text="Save", command=save_to_excel,  bg='#E8DAEF', width=10, height=1)
        save_button.pack()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"employees_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")

    def insert4():
        try:
            tree.delete(*tree.get_children())  # удаление устаревших данных

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM employees")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row,)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def add_string():
        id1 = entry.get()
        fio = entry2.get()
        job_titles = entry3.get()
        number = entry4.get()
        email = entry5.get()
        id2 = entry6.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO employees (\"ID_персонала\", ФИО, Должность, Номер_телефона, Почта, "
                           "\"ID_отдела\") VALUES (%s, %s, %s, %s, %s, %s)",
                           (id1, fio, job_titles, number, email, id2))
            conn.commit()

            insert4()  # Обновить вывод таблицы после добавления строки
            entry.delete(0, tk.END)
            entry2.delete(0, tk.END)
            entry3.delete(0, tk.END)
            entry4.delete(0, tk.END)
            entry5.delete(0, tk.END)
            entry6.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root3.bind('<Return>', on_enter)

    def delete_selected_employee():
        try:
            selected_item = tree.selection()

            for item in selected_item:
                values = tree.item(item, 'values')
                ID_employee = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                try:
                    # Удалить из medical_information
                    cursor.execute("DELETE FROM medical_information WHERE \"ID_персонала\" = %s", (ID_employee,))

                    # Удалить из accounting
                    cursor.execute("DELETE FROM accounting WHERE \"ID_персонала\" = %s", (ID_employee,))

                    # Удалить из labor_activity
                    cursor.execute("DELETE FROM labor_activity WHERE \"ID_персонала\" = %s", (ID_employee,))

                    # Удалить из employees последним
                    cursor.execute("DELETE FROM employees WHERE \"ID_персонала\" = %s", (ID_employee,))

                    # Применить все изменения
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    error_msg = f"Ошибка при удалении данных из одной из таблиц: {e}"
                    print(error_msg)
                    messagebox.showerror("Ошибка", error_msg)
                    return

            insert4()  # Обновить вывод таблицы после удаления строк
        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при выполнении удаления из базы данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_selected_employee()
        else:
            print("Action canceled.")

    def update_record():
        try:
            selected_item = tree.selection()

            if not selected_item:
                messagebox.showinfo("Информация", "Пожалуйста, выберите строку для изменения.")
                return

            id1 = entry.get()
            fio = entry2.get()
            job_title = entry3.get()
            number = entry4.get()
            email = entry5.get()
            id_department = entry6.get()

            if not id1 or not fio or not job_title or not number or not email or not id_department:
                messagebox.showinfo("Информация", "Пожалуйста, заполните все поля для обновления.")
                return

            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                employee_id = values[0]

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для обновления строки по определенному идентификатору
                cursor.execute(
                    "UPDATE employees SET \"ID_персонала\" = %s, ФИО = %s, Должность = %s, Номер_телефона = %s, Почта = %s, \"ID_отдела\" = %s WHERE \"ID_персонала\" = %s",
                    (id1, fio, job_title, number, email, id_department, employee_id))

            conn.commit()
            insert4()  # Обновить вывод таблицы после обновления строки
            print("Запись успешно обновлена в базе данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении строки в базе данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        # Очистка полей ввода после обновления
        entry.delete(0, tk.END)
        entry2.delete(0, tk.END)
        entry3.delete(0, tk.END)
        entry4.delete(0, tk.END)
        entry5.delete(0, tk.END)
        entry6.delete(0, tk.END)

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле
            entry3.delete(0, tk.END)
            entry3.insert(0, item_values[2])  # Пример: ID отдела в первое поле
            entry4.delete(0, tk.END)
            entry4.insert(0, item_values[3])  # Пример: Название отдела во второе поле
            entry5.delete(0, tk.END)
            entry5.insert(0, item_values[4])  # Пример: ID отдела в первое поле
            entry6.delete(0, tk.END)
            entry6.insert(0, item_values[5])  # Пример: Название отдела во второе поле

    tree.bind("<Double-1>", on_double_click)

    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    def delete1():
        # очищение первой строки
        entry.delete(0, tk.END)

    def delete2():
        # очищение второй строки
        entry2.delete(0, tk.END)

    def delete3():
        # очищение первой строки
        entry3.delete(0, tk.END)

    def delete4():
        # очищение второй строки
        entry4.delete(0, tk.END)

    def delete5():
        # очищение первой строки
        entry5.delete(0, tk.END)

    def delete6():
        # очищение второй строки
        entry6.delete(0, tk.END)

    def delete7():
        # очищение строки запроса
        sort_entry.delete(0, tk.END)

    entry_frame10 = tk.Frame(root3, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=60, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=25, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_record,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame11 = tk.Frame(root3, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert4,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=60, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=25, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=25, pady=2)
    # Создание кнопки "сортировать"
    btn7 = tk.Button(entry_frame11, text="Sort", padx=5, pady=1, width=10, height=1, command=sort_by_department,
                     bg='#E8DAEF')
    btn7.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame12 = tk.Frame(root3, background="#D2B4DE")
    entry_frame12.pack(side='right')
    # Создание кнопок "x"
    btn1 = tk.Button(entry_frame12, text="х", command=delete1,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, pady=1)
    btn2 = tk.Button(entry_frame12, text="х", command=delete2,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, pady=1)
    btn3 = tk.Button(entry_frame12, text="х", command=delete3,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, pady=1)
    btn4 = tk.Button(entry_frame12, text="х", command=delete4,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, pady=1)
    btn5 = tk.Button(entry_frame12, text="х", command=delete5,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, pady=1)
    btn6 = tk.Button(entry_frame12, text="х", command=delete6,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, pady=1)
    btn7 = tk.Button(entry_frame12, text="х", command=delete7,
                     bg='#E8DAEF')
    btn7.pack(side=tk.TOP, pady=1)

    entry_frame1 = tk.Frame(root3, background="#D2B4DE")
    entry_frame1.pack()

    entry_label = tk.Label(entry_frame1, text="Введите id сотрудника:", width=30, height=2, background="#D2B4DE",
                           anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame1, width=30)  # Создаем поле ввода шириной 30 символов
    entry.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame2 = tk.Frame(root3, background="#D2B4DE")
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите ФИО сотрудника:", width=30, height=1, background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=30)  # Создаем поле ввода шириной 30 символов
    entry2.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame3 = tk.Frame(root3, background="#D2B4DE")
    entry_frame3.pack()

    entry_label3 = tk.Label(entry_frame3, text="Введите должность сотрудника:", width=30, height=2,
                            background="#D2B4DE", anchor='w')
    entry_label3.pack(side='left')

    entry3 = tk.Entry(entry_frame3, width=30)  # Создаем поле ввода шириной 30 символов
    entry3.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame4 = tk.Frame(root3, background="#D2B4DE")
    entry_frame4.pack()

    entry_label4 = tk.Label(entry_frame4, text="Введите номер телефона сотрудника:", width=30, height=1,
                            background="#D2B4DE", anchor='w')
    entry_label4.pack(side='left')

    entry4 = tk.Entry(entry_frame4, width=30)  # Создаем поле ввода шириной 30 символов
    entry4.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame5 = tk.Frame(root3, background="#D2B4DE")
    entry_frame5.pack()

    entry_label5 = tk.Label(entry_frame5, text="Введите почту сотрудника:", width=30, height=2, background="#D2B4DE",
                            anchor='w')
    entry_label5.pack(side='left')

    entry5 = tk.Entry(entry_frame5, width=30)  # Создаем поле ввода шириной 30 символов
    entry5.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame6 = tk.Frame(root3, background="#D2B4DE")
    entry_frame6.pack()

    entry_label6 = tk.Label(entry_frame6, text="Введите id отдела:", width=30, height=1, background="#D2B4DE",
                            anchor='w')
    entry_label6.pack(side='left')

    entry6 = tk.Entry(entry_frame6, width=30)  # Создаем поле ввода шириной 30 символов
    entry6.pack(side='left')  # Размещаем поле ввода на frame1

    # Создание поля для сортировки

    sort_entry_frame = tk.Frame(root3, background="#D2B4DE")
    sort_entry_frame.pack()

    sort_label = tk.Label(sort_entry_frame, text="Введите ID отдела для сортировки:", width=30, height=2,
                          background="#D2B4DE", anchor='w')
    sort_label.pack(side='left')

    sort_entry = tk.Entry(sort_entry_frame, width=30, background="light gray")
    sort_entry.pack(side='left')

    return entry, entry2, entry3, entry4, entry5, entry6, sort_entry # Возвращаем созданное поле ввода


def labor_activity():
    # Создаем графический интерфейс
    root4 = tk.Tk()
    root4.title("Labor activity")
    root4.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root4.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root4)
    tree["columns"] = ('ID employees', 'Apply date', 'Time work', 'Fire date')
    tree.heading('#0', text='№')
    tree.heading('ID employees', text='ID employees')
    tree.heading('Apply date', text='Apply date')
    tree.heading('Time work', text='Time work')
    tree.heading('Fire date', text='Fire date')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID employees", width=100, anchor=tk.CENTER)
    tree.column("Apply date", width=100, anchor=tk.CENTER)
    tree.column("Time work", width=100, anchor=tk.CENTER)
    tree.column("Fire date", width=100, anchor=tk.CENTER)

    tree.pack()

    try:
        tree.delete(*tree.get_children())  # удаление устаревших данных

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM labor_activity")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row, )

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")

    def sort_by_labor():
        selected_department_id = entry5.get()

        sorted_employees = []
        for child in tree.get_children():
            values = tree.item(child, 'values')
            if values and values[1] == selected_department_id:
                sorted_employees.append(values)

        # Создание нового окна для вывода отсортированных данных
        result_window = tk.Toplevel(root4)
        result_window.title("Sorted Employees by Labor")
        result_window.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
        result_window.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

        result_tree = ttk.Treeview(result_window)
        result_tree["columns"] = ('ID employees', 'Date_us', 'Time fire', 'Date_uv')

        # Создание колонок таблицы
        for i, column in enumerate(('ID employees', 'Date_us', 'Time fire', 'Date_uv')):
            result_tree.heading(i, text=column)
            result_tree.column(i, width=180, minwidth=100, anchor=tk.CENTER, stretch=True)

        for idx, labor in enumerate(sorted_employees):
            result_tree.insert("", "end", text=str(idx), values=labor)

        result_tree.pack()

        # Функция сохранения результатов сортировки
        def save_to_excel():
            wb = Workbook()
            ws = wb.active
            for i, column in enumerate(('ID employees', 'Date_us', 'Time fire', 'Date_uv')):
                ws.cell(row=1, column=i + 1, value=column)

            for idx, employee in enumerate(sorted_employees):
                for i, value in enumerate(employee):
                    ws.cell(row=idx + 2, column=i + 1, value=value)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"labor_data_{timestamp}.xlsx")

            wb.save(save_path)
            print("Данные успешно сохранены в файле labor_data.xlsx. на рабочий стол")

        # Создание кнопки "Сохранить"
        save_button = tk.Button(result_window, text="Save", command=save_to_excel,  bg='#E8DAEF', width=10, height=1)
        save_button.pack()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"labor_activity_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")



    def insert5():
        try:
            tree.delete(*tree.get_children())  # удаление устаревших данных

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM labor_activity")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row,)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def add_string():
        id1 = entry.get()
        date_us = entry2.get()
        time_job = entry3.get()
        date_uv = entry4.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO labor_activity (\"ID_персонала\", Дата_устройства, "
                           "Количество_времени_работы, Дата_увольнения) VALUES (%s, %s, %s, %s)",
                           (id1, date_us, time_job, date_uv))
            conn.commit()

            insert5()  # Обновить вывод таблицы после добавления строки
            entry.delete(0, tk.END)
            entry2.delete(0, tk.END)
            entry3.delete(0, tk.END)
            entry4.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root4.bind('<Return>', on_enter)

    def delete_record():
        try:
            selected_item = tree.selection()
            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                Labor = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

            # Выполнение SQL-запроса для удаления строки по определенному идентификатору
            cursor.execute("DELETE FROM labor_activity WHERE \"ID_персонала\" = %s", (Labor,))

            conn.commit()
            insert5()  # Обновить вывод таблицы после добавления строки
            print("Строка успешно удалена из базы данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при удалении строки из базы данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_record()
        else:
            print("Action canceled.")

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле
            entry3.delete(0, tk.END)
            entry3.insert(0, item_values[2])  # Пример: ID отдела в первое поле
            entry4.delete(0, tk.END)
            entry4.insert(0, item_values[3])  # Пример: Название отдела во второе поле


    tree.bind("<Double-1>", on_double_click)

    def update_record():
        try:
            selected_item = tree.selection()

            if not selected_item:
                messagebox.showinfo("Информация", "Пожалуйста, выберите строку для изменения.")
                return

            id1 = entry.get()
            date_us = entry2.get()
            time_job = entry3.get()
            date_uv = entry4.get()

            if not id1 or not date_us or not time_job or not date_uv:
                messagebox.showinfo("Информация", "Пожалуйста, заполните все поля для обновления.")
                return

            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                employee_id = values[0]

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для обновления строки по определенному идентификатору
                cursor.execute(
                    "UPDATE labor_activity SET \"ID_персонала\" = %s, Дата_устройства = %s, Количество_времени_работы = %s, Дата_увольнения = %s WHERE \"ID_персонала\" = %s",
                    (id1, date_us, time_job, date_uv, employee_id))

            conn.commit()
            insert5()  # Обновить вывод таблицы после обновления строки
            print("Запись успешно обновлена в базе данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении строки в базе данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        # Очистка полей ввода после обновления
        entry.delete(0, tk.END)
        entry2.delete(0, tk.END)
        entry3.delete(0, tk.END)
        entry4.delete(0, tk.END)

    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    def delete1():
        # очищение первой строки
        entry.delete(0, tk.END)

    def delete2():
        # очищение второй строки
        entry2.delete(0, tk.END)

    def delete3():
        # очищение первой строки
        entry3.delete(0, tk.END)

    def delete4():
        # очищение второй строки
        entry4.delete(0, tk.END)

    def delete5():
        # очищение 5ой строки
        entry5.delete(0, tk.END)

    entry_frame10 = tk.Frame(root4, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_record,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame12 = tk.Frame(root4, background="#D2B4DE")
    entry_frame12.pack(side='right')

    entry_frame11 = tk.Frame(root4, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert5,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=25, pady=2)

    btn7 = tk.Button(entry_frame11, text="Sort", padx=5, pady=1, width=10, height=1, command=sort_by_labor,
                     bg='#E8DAEF')
    btn7.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame12 = tk.Frame(root4, background="#D2B4DE")
    entry_frame12.pack(side='right')
    # Создание кнопок "x"
    btn1 = tk.Button(entry_frame12, text="х", command=delete1,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, pady=1, padx=4)
    btn2 = tk.Button(entry_frame12, text="х", command=delete2,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, pady=1)
    btn3 = tk.Button(entry_frame12, text="х", command=delete3,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, pady=1)
    btn4 = tk.Button(entry_frame12, text="х", command=delete4,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, pady=1)
    btn5 = tk.Button(entry_frame12, text="х", command=delete5,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, pady=1)

    entry_frame1 = tk.Frame(root4, background="#D2B4DE")
    entry_frame1.pack()

    entry_label = tk.Label(entry_frame1, text="Введите id сотрудника:", width=20, height=1,  background="#D2B4DE",
                           anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame1, width=30)  # Создаем поле ввода шириной 30 символов
    entry.pack(side='left')  # Размещаем поле ввода на frame1

    entry_frame2 = tk.Frame(root4, background="#D2B4DE")
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите дату устройства:", width=20, height=2,  background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=30)  # Создаем поле ввода шириной 30 символов
    entry2.pack(side='left')  # Размещаем поле ввода на frame2

    entry_frame3 = tk.Frame(root4, background="#D2B4DE")
    entry_frame3.pack()

    entry_label3 = tk.Label(entry_frame3, text="Введите прод. работы:", width=20, height=1,  background="#D2B4DE",
                            anchor='w')
    entry_label3.pack(side='left')

    entry3 = tk.Entry(entry_frame3, width=30)  # Создаем поле ввода шириной 30 символов
    entry3.pack(side='left')  # Размещаем поле ввода на frame3

    entry_frame4 = tk.Frame(root4, background="#D2B4DE")
    entry_frame4.pack()

    entry_label4 = tk.Label(entry_frame4, text="Введите дату увольнения:", width=20, height=2,  background="#D2B4DE",
                            anchor='w')
    entry_label4.pack(side='left')

    entry4 = tk.Entry(entry_frame4, width=30)  # Создаем поле ввода шириной 30 символов
    entry4.pack(side='left')  # Размещаем поле ввода на frame4

    entry_frame5 = tk.Frame(root4, background="#D2B4DE")
    entry_frame5.pack()

    entry_label5 = tk.Label(entry_frame5, text="Введите дату для сорт-ки:", width=20, height=1, background="#D2B4DE",
                            anchor='w')
    entry_label5.pack(side='left')

    entry5 = tk.Entry(entry_frame5, width=30, bg='light gray')  # Создаем поле ввода шириной 30 символов
    entry5.pack(side='left')  # Размещаем поле ввода на frame5

    return entry, entry2, entry3, entry4, entry5  # Возвращаем созданное поле ввода

    # root4.mainloop()
def medical_information():
    # Создаем графический интерфейс
    root5 = tk.Tk()
    root5.title("Medical_information")
    root5.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root5.configure(bg='#D2B4DE')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root5)
    tree["columns"] = ('ID employees', 'Medical statistic')
    tree.heading('#0', text='№')
    tree.heading('ID employees', text='ID employees')
    tree.heading('Medical statistic', text='Medical statistic')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=40)
    tree.column("ID employees", width=100, anchor=tk.CENTER)
    tree.column("Medical statistic", width=160, anchor=tk.CENTER)

    tree.pack()

    try:
        tree.delete(*tree.get_children())  # удаление устаревших данных

        cursor = conn.cursor()
        cursor.execute("SELECT * FROM medical_information")
        rows = cursor.fetchall()

        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row, )

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"medical_informetion_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("Данные успешно сохранены в файл Excel!")
        except Exception as e:
            print(f"Ошибка при сохранении данных в Excel: {e}")


    # Создание кнопок
    def insert6():
        try:
            tree.delete(*tree.get_children())  # удаление устаревших данных

            cursor = conn.cursor()
            cursor.execute("SELECT * FROM medical_information")
            rows = cursor.fetchall()

            for i, row in enumerate(rows):
                tree.insert("", "end", text=str(i), values=row,)

        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")

    def add_string():
        id1 = entry.get()
        med_book = entry2.get()

        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO medical_information (\"ID_персонала\", Данные_медкниги) VALUES (%s, %s)",
                           (id1, med_book))
            conn.commit()

            insert6()  # Обновить вывод таблицы после добавления строки
            entry.delete(0, tk.END)
            entry2.delete(0, tk.END)

        except Exception as e:
            # Если произошла ошибка, откатываем транзакцию
            conn.rollback()

            error_msg = f"Ошибка при загрузке данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def on_enter(event):
        # Поместите здесь ваше действие
        add_string()

    # Привязываем клавишу Enter к функции on_enter
    root5.bind('<Return>', on_enter)

    def delete_record():
        try:
            selected_item = tree.selection()
            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                Medical = values[0]

                cursor = conn.cursor()

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                cursor.execute("DELETE FROM medical_information WHERE \"ID_персонала\" = %s", (Medical,))

            conn.commit()
            insert6()  # Обновить вывод таблицы после добавления строки
            print("Строка успешно удалена из базы данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при удалении строки из базы данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

    def confirm_action():
        result = messagebox.askyesno("Confirmation", "Вы уверены, что хотите очистить?")
        if result:
            # Put your action here
            delete_record()
        else:
            print("Action canceled.")

    def on_double_click(event):
        selected_item = tree.selection()[0]  # Получаем ID выбранной строки
        item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
        if item_values:
            entry.delete(0, tk.END)
            entry.insert(0, item_values[0])  # Пример: ID отдела в первое поле
            entry2.delete(0, tk.END)
            entry2.insert(0, item_values[1])  # Пример: Название отдела во второе поле

    tree.bind("<Double-1>", on_double_click)

    def update_record():
        try:
            selected_item = tree.selection()

            if not selected_item:
                messagebox.showinfo("Информация", "Пожалуйста, выберите строку для изменения.")
                return

            id1 = entry.get()
            med_book = entry2.get()

            if not id1 or not med_book:
                messagebox.showinfo("Информация", "Пожалуйста, заполните все поля для обновления.")
                return

            cursor = conn.cursor()

            for item in selected_item:
                values = tree.item(item, 'values')
                employee_id = values[0]

                # Начало транзакции
                cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для обновления строки по определенному идентификатору
                cursor.execute(
                    "UPDATE medical_information SET \"ID_персонала\" = %s, Данные_медкниги = %s WHERE \"ID_персонала\" = %s",
                    (id1, med_book, employee_id))

            conn.commit()
            insert6()  # Обновить вывод таблицы после обновления строки
            print("Запись успешно обновлена в базе данных.")

        except Exception as e:
            conn.rollback()
            error_msg = f"Ошибка при обновлении строки в базе данных: {e}"
            print(error_msg)
            messagebox.showerror("Ошибка", error_msg)

        # Очистка полей ввода после обновления
        entry.delete(0, tk.END)
        entry2.delete(0, tk.END)


    def delete():
        # удаление устаревших данных
        tree.delete(*tree.get_children())

    def delete1():
        # очищение первой строки
        entry.delete(0, tk.END)

    def delete2():
        # очищение второй строки
        entry2.delete(0, tk.END)


    entry_frame10 = tk.Frame(root5, background="#D2B4DE")
    entry_frame10.pack(side='right')
    # Создание кнопки "добавить строку"
    btn2 = tk.Button(entry_frame10, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить строку"
    btn4 = tk.Button(entry_frame10, text="Clear String", padx=5, pady=1, width=10, height=1, command=confirm_action,
                     bg='#E8DAEF')
    btn4.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "изменить"
    btn5 = tk.Button(entry_frame10, text="Сhange", padx=5, pady=1, width=10, height=1, command=update_record,
                     bg='#E8DAEF')
    btn5.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame11 = tk.Frame(root5, background="#D2B4DE")
    entry_frame11.pack(side='right')
    # Создание кнопки "импорт"
    btn1 = tk.Button(entry_frame11, text="Import Table", padx=5, pady=1, width=10, height=1, command=insert6,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "очистить"
    btn3 = tk.Button(entry_frame11, text="Clear Page", padx=5, pady=1, width=10, height=1, command=delete,
                     bg='#E8DAEF')
    btn3.pack(side=tk.TOP, padx=10, pady=2)
    # Создание кнопки "сохранить"
    btn6 = tk.Button(entry_frame11, text="Save", padx=5, pady=1, width=10, height=1, command=save_to_excel,
                     bg='#E8DAEF')
    btn6.pack(side=tk.TOP, padx=25, pady=2)

    entry_frame12 = tk.Frame(root5, background="#D2B4DE")
    entry_frame12.pack(side='right')
    # Создание кнопок "x"
    btn1 = tk.Button(entry_frame12, text="х", command=delete1,
                     bg='#E8DAEF')
    btn1.pack(side=tk.TOP, pady=1, padx=4)
    btn2 = tk.Button(entry_frame12, text="х", command=delete2,
                     bg='#E8DAEF')
    btn2.pack(side=tk.TOP, pady=11)

    entry_frame1 = tk.Frame(root5, pady=2,  background="#D2B4DE")
    entry_frame1.pack()

    entry_label = tk.Label(entry_frame1, text="Введите id сотрудника:", width=22, height=2,  background="#D2B4DE",
                           anchor='w')
    entry_label.pack(side='left')

    entry = tk.Entry(entry_frame1, width=22)
    entry.pack(side='left')

    entry_frame2 = tk.Frame(root5,  background="#D2B4DE")
    entry_frame2.pack()

    entry_label2 = tk.Label(entry_frame2, text="Введите данные мед книги:", width=22, height=2,  background="#D2B4DE",
                            anchor='w')
    entry_label2.pack(side='left')

    entry2 = tk.Entry(entry_frame2, width=22)
    entry2.pack(side='left')


    root5.mainloop()


def documentation ():
    boom1 = tk.Tk()
    boom1.title("Data Base 📊")
    boom1.geometry("700x450")
    boom1.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    boom1.configure(bg='#D0ECE7')  # Задание цвета фона окна в формате #RRGGBB

    label = tk.Label(boom1, text="", font=('Arial', 9, 'normal'), background="#D0ECE7", foreground="white")
    label.pack(pady=5)

    # Создание поля  текстом
    text = tk.Text(boom1)
    text.pack()
    text.tag_configure("custom_font", font=("Helvetica", 10, "bold"))
    text.insert(tk.END, "# Документация по использованию приложения с базой данных Dashapupu Corporation\n\n"
        "## 1. Введение\n"
        "Данная документация предоставляет подробные инструкции по использованию\n"
        "приложения Dashapupu Corporation, которое предназначено для\n"
        "взаимодействия с базой данных сотрудников. В этом документе содержится\n"
        "информация о начальной настройке, выполнении основных операций и\n"
        "рекомендациях по эффективному использованию приложения.\n\n"
        "## 2. Начало работы\n\n"
        "### 2.1. Регистрация и вход в приложение\n"
        "Для начала работы с приложением необходимо выполнить следующие шаги:\n\n"
        "1. Загрузка и установка приложения: Загрузите установочный файл\n"
        "приложения с официального сайта Dashapupu Corporation и следуйте\n"
        "инструкциям по установке.\n"
        "2. Регистрация нового пользователя:\n"
        "   - Запустите приложение и нажмите кнопку Registration на главном экране.\n"
        "   - Введите необходимую информацию, включая имя и пароль.\n"
        "   - Подтвердите регистрацию, следуя инструкциям на экране.\n"
        "3. Вход в приложение:\n"
        "   - На главном экране нажмите кнопку Login.\n"
        "   - Введите ваше имя пользователя и пароль, затем нажмите кнопку Войти\n"
        "     для доступа к функционалу приложения.\n\n"
        "### 2.2. Подключение к базе данных\n"
        "Приложение автоматически подключается к базе данных при успешном\n"
        "входе в систему. В случае возникновения проблем с подключением\n"
        "убедитесь, что у вас есть стабильное интернет-соединение. Если проблема\n"
        "не решается, обратитесь к системному администратору для получения помощи.\n\n"
        "## 3. Основные операции\n\n"
        "### 3.1. Просмотр данных\n"
        "Для доступа к различным разделам данных используйте основное меню\n"
        "приложения. Основное меню включает следующие разделы:\n\n"
        "- Employees - управление данными сотрудников.\n"
        "- Labor activity - информация о трудовой деятельности.\n"
        "- Accounting - учет финансовых данных.\n"
        "- Medical information - медицинская информация сотрудников.\n"
        "- Department - данные о подразделениях.\n"
        "- Statistics - статистические данные и отчеты.\n\n"
        "Для просмотра конкретных записей выберите соответствующий раздел.\n"
        "Например, чтобы просмотреть данные о сотрудниках, выберите раздел\n"
        "Employees.\n\n"
        "### 3.2. Добавление данных\n"
        "Для добавления новых записей в базу данных выполните следующие шаги:\n\n"
        "1. Перейдите в соответствующий раздел, например, Employees.\n"
        "2. Заполните необходимые поля в форме (ID сотрудника, ФИО, должность,\n"
        "   номер телефона, email, ID отдела).\n"
        "3. Нажмите кнопку Add String для сохранения новой записи в базу данных.\n\n"
        "### 3.3. Редактирование и удаление данных\n"
        "Для редактирования или удаления существующих записей выполните следующие шаги:\n\n"
        "- Редактирование записи:\n"
        "  - Выберите запись, которую необходимо изменить.\n"
        "  - Внесите изменения в соответствующие поля.\n"
        "  - Нажмите кнопку Change для сохранения изменений.\n\n"
        "- Удаление записи:\n"
        "  - Выберите запись, которую необходимо удалить.\n"
        "  - Подтвердите удаление, следуя инструкциям на экране.\n\n"
        "## 4. Рекомендации по использованию\n\n"
        "### 4.1. Безопасность\n"
        "- Никогда не передавайте свои учетные данные третьим лицам.\n"
        "- Используйте надежные пароли для защиты доступа к приложению и базе данных.\n"
        "- Регулярно обновляйте пароли и следите за безопасностью вашего аккаунта.\n\n"
        "### 4.2. Резервное копирование\n"
        "- Регулярно создавайте резервные копии данных для предотвращения\n"
        "  потери информации.\n"
        "- Обратитесь к администратору базы данных для получения дополнительной\n"
        "  информации о процедурах резервного копирования и восстановления данных.\n\n"
        "## 5. Заключение\n"
        "Приложение Dashapupu Corporation предлагает мощные и интуитивно\n"
        "понятные инструменты для управления данными сотрудников. Соблюдайте\n"
        "рекомендации по безопасности и резервному копированию, чтобы обеспечить\n"
        "надежную и эффективную работу приложения. Следуйте инструкциям в данной\n"
        "документации для оптимального использования всех возможностей,\n"
        "предоставляемых приложением.", "custom_font")

    boom1.mainloop()

def data():

    def close():
        boom.destroy()


    boom = tk.Tk()
    boom.title("Data Base 📊")
    boom.geometry("1080x550")
    boom.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    boom.configure(bg='#EBDEF0')  # Задание цвета фона окна в формате #RRGGBB

    label = tk.Label(boom, text="", font=('Arial', 9, 'normal'), background="#EBDEF0", foreground="white")
    label.pack(pady=5)

    label1 = tk.Label(boom, text="Data Base dashapupu corporation", font=('Arial', 12, 'normal'), background="#EBDEF0", foreground="gray")
    label1.pack(pady=5)

    colon_l = tk.Frame(boom, background="#EBDEF0", padx=50, pady=1)
    colon_l.pack(side='left')

    # Создание кнопок
    Employees = tk.Button(colon_l, text="Employees 👥", command=employees, width=20, height=2, background="#AF7AC5",
                          foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Employees.pack(pady=6)

    Labor_activity = tk.Button(colon_l, text="Labor activity ⏱️", command=labor_activity, width=20, height=2,
                               background="#AF7AC5", foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'),
                               cursor="hand2")
    Labor_activity.pack(pady=6)

    Accounting = tk.Button(colon_l, text="Accounting 💰", command=accounting, width=20, height=2, background="#AF7AC5",
                           foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Accounting.pack(pady=6)

    colon_r = tk.Frame(boom, background="#EBDEF0", padx=50, pady=1)
    colon_r.pack(side='right')

    Medical_information = tk.Button(colon_r, text="Medical information 💊", command=medical_information, width=20, height=2,
                                    background="#AF7AC5", foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Medical_information.pack(pady=6)

    Department = tk.Button(colon_r, text="Department 📂", command=department, width=20, height=2, background="#AF7AC5",
                           foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Department.pack(pady=6)

    Department_statistics = tk.Button(colon_r, text="Statistics 📊", command=department_statistics, width=20, height=2,
                                      background="#AF7AC5", foreground="white", borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Department_statistics.pack(pady=6)

    Exit_frame = tk.Frame(boom, background="#EBDEF0")
    Exit_frame.pack(side='bottom', pady=10)

    # Back = tk.Button(Exit_frame, text="Back", command=back, width=10, height=1, background="#AF7AC5",
    #                  foreground="white", borderwidth=8, font=('Verdana', 20, 'normal'), cursor="hand2")
    # Back.pack()

    Exit = tk.Button(Exit_frame, text="Exit", command=close, width=10, height=1, background="#AF7AC5",
                     foreground="white", borderwidth=8, font=('Verdana', 20, 'normal'), cursor="hand2")
    Exit.pack()


def main_menu():

    def close():
        menu.destroy()

    menu = tk.Tk()
    menu.title("Menu")
    menu.geometry("500x500")
    menu.resizable(False, False)
    menu.configure(bg="#D0ECE7")


    label = tk.Label(menu, text="", font=('Arial', 11, 'normal'), background="#D0ECE7", foreground="gray")
    label.pack(pady=20)


        # Создание кнопок
    Data = tkinter.Button(menu, text="Data Base",command=data, width=13, height=2,background="#27AE60", foreground="white",
                                          borderwidth="8",font=('Verdana', 20, 'normal'), cursor="hand2" )
    Data.pack(pady=10)

    Documentation = tk.Button(menu, text="Documentation", command=documentation, width=13, height=2, background="#27AE60", foreground="white",
                                     borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Documentation.pack(pady=10)

    Exit = tk.Button(menu, text="Exit", command=close, width=13, height=2, background="#27AE60", foreground="white",
                                     borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2")
    Exit.pack(pady=10)

    label = tk.Label(menu, text="        dashapupu corporation©️", font=('Arial', 14, 'normal'), background="#D0ECE7")
    label.pack(pady=28)

    # menu.mainloop()


# Функция для сохранения пользователей в файл
def save_users():
    with open('users.pickle', 'wb') as file:
        pickle.dump(users, file)

# Загрузка пользователей из файла или создание нового словаря, если файла нет
try:
    with open('users.pickle', 'rb') as file:
        users = pickle.load(file)
except FileNotFoundError:
    users = {"admin": "password"}

def reg():
    def register():
        username = entry_username.get()
        password = entry_password.get()

        # Регистрация нового пользователя
        if username not in users:
            users[username] = password
            save_users()  # Сохраняем пользователей после добавления нового
            label_status.config(text="Пользователь зарегистрирован.")
        else:
            label_status.config(text="Пользователь уже существует.")

    root8 = tk.Tk()
    root8.title("Регистрация")
    root8.geometry("200x170")
    root8.configure(bg="#D6EAF8")

    label_username = tk.Label(root8, text="", bg="#D6EAF8")
    label_username.pack()

    label_username = tk.Label(root8, text="Имя пользователя:", bg="#D6EAF8")
    label_username.pack()
    entry_username = tk.Entry(root8)
    entry_username.pack()

    label_password = tk.Label(root8, text="Пароль:", bg="#D6EAF8")
    label_password.pack()
    entry_password = tk.Entry(root8, show="*")
    entry_password.pack()

    button_register = tk.Button(root8, text="Зарегистрироваться", command=register, bg="#3498DB", foreground="white",
                                borderwidth="2")
    button_register.pack(pady=10)

    label_status = tk.Label(root8, text="", bg="#D6EAF8")
    label_status.pack()

    root8.mainloop()


def log():
    def login():
        username = entry_username.get()
        password = entry_password.get()

        # Проверка логина и пароля
        if username in users and users[username] == password:
            main_menu()
            root7.destroy()
            entrance.destroy()
        else:
            label_status.config(text="Неверный логин или пароль.")

    root7 = tk.Tk()
    root7.title("Авторизация")
    root7.geometry("200x170")
    root7.configure(bg="#D6EAF8")

    label_username = tk.Label(root7, text="", bg="#D6EAF8")
    label_username.pack()

    label_username = tk.Label(root7, text="Имя пользователя:", bg="#D6EAF8")
    label_username.pack()
    entry_username = tk.Entry(root7)
    entry_username.pack()

    label_password = tk.Label(root7, text="Пароль:", bg="#D6EAF8")
    label_password.pack()
    entry_password = tk.Entry(root7, show="*")
    entry_password.pack()

    button_login = tk.Button(root7, text="Войти", command=login, bg="#3498DB", foreground="white", borderwidth="2")
    button_login.pack(pady=10)

    label_status = tk.Label(root7, text="", bg="#D6EAF8")
    label_status.pack()

    root7.mainloop()

def close_add_ed():
    entrance.destroy()

# Начало работы, то, что с самого начал выводится на экран:


# Создание главного окна

entrance = tk.Tk()
entrance.title("Entrance")
entrance.geometry("300x330")
entrance.resizable(False, False)
entrance.configure(bg="#D6EAF8")

label = tk.Label(entrance, text="   dashapupu corporation©️", font=('Arial', 14, 'normal'), background="#D6EAF8")
label.pack(pady=10)

login1=tk.Button(entrance, text ="Login", width=10, height=1, background="#3498DB", foreground="white",
                 borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2", command=log)
login1.pack(pady=10)

register1=tk.Button(entrance, text ="Registration", width=10, height=1, background="#3498DB", foreground="white",
                    borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2", command=reg)
register1.pack(pady=10)

exit1=tk.Button(entrance, text ="Exit", width=10, height=1, background="#3498DB", foreground="white",
                borderwidth="8", font=('Verdana', 20, 'normal'), cursor="hand2", command=close_add_ed)
exit1.pack(pady=10)

entrance.mainloop()
