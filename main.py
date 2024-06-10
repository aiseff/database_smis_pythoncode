import pickle
import tkinter as tk
from tkinter import ttk
import psycopg2
from tkinter import messagebox
from psycopg2 import Error
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
# Подключение к базе данных PostgreSQL
conn = psycopg2.connect(
    dbname="smis",
    user="postgres",
    password="admin",
    host="localhost",
    port="5432"
)
cursor = conn.cursor()

def product():

    def clear_all_entrys():
        entry_id.delete(0, tk.END)
        entry_name.delete(0, tk.END)
        entry_price1.delete(0, tk.END)
        # entry_remain.delete(0, tk.END)
        # entry_sold.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Products")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Product id','Product name','Product price','Product remain','Product sold','Product spent','Product gain')
    tree.heading('#0', text='Number')
    tree.heading('Product id', text='Product id')
    tree.heading('Product name', text='Product name')
    tree.heading('Product price', text='Product price')
    tree.heading('Product remain', text='Product remain')
    tree.heading('Product sold', text='Product sold')
    tree.heading('Product spent', text='Product spent')
    tree.heading('Product gain', text='Product gain')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=60)
    tree.column("Product id", width=70, anchor=tk.CENTER)
    tree.column("Product name", width=220, anchor=tk.CENTER)
    tree.column("Product price", width=80, anchor=tk.CENTER)
    tree.column("Product remain", width=95, anchor=tk.CENTER)
    tree.column("Product sold", width=85, anchor=tk.CENTER)
    tree.column("Product spent", width=90, anchor=tk.CENTER)
    tree.column("Product gain", width=80, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM product")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM product")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM product")
        rows = cursor.fetchall()

#Sorting
    def sort_by_id():
        try:
            selected_product_id = int(entry_sort.get())
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = ('Product id','Product name','Product price','Product remain','Product sold','Product spent','Product gain')
            result_tree.heading('#0', text='№')
            result_tree.heading('Product id', text='Product id')
            result_tree.heading('Product name', text='Product name')
            result_tree.heading('Product price', text='Product price')
            result_tree.heading('Product remain', text='Product remain')
            result_tree.heading('Product sold', text='Product sold')
            result_tree.heading('Product spent', text='Product spent')
            result_tree.heading('Product gain', text='Product gain')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Product id", width=100, anchor=tk.CENTER)
            result_tree.column("Product name", width=220, anchor=tk.CENTER)
            result_tree.column("Product price", width=100, anchor=tk.CENTER)
            result_tree.column("Product remain", width=100, anchor=tk.CENTER)
            result_tree.column("Product sold", width=100, anchor=tk.CENTER)
            result_tree.column("Product spent", width=100, anchor=tk.CENTER)
            result_tree.column("Product gain", width=100, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM product WHERE product_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        def save_to_excel(selected_product_id):
            cursor.execute("SELECT * FROM product WHERE product_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            ws.append(['Product id', 'Product name', 'Product price', 'Product remain', 'Product sold','Product spent','Product gain'])
            for row in rows:
                ws.append(row)
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            wb.save(save_path)
            print(f"The data has been successfully saved in the file {save_path} on the desktop.")

        # Создание кнопки "Save"
        button_save = tk.Button(result_window, text='Save', width=10, height=1, command=lambda: save_to_excel(selected_product_id))
        button_save.pack()

#Saving the table
    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()
        def add_string():
            id = int(entry_id.get())
            name = entry_name.get()
            price1 = int(entry_price1.get())
            remain = 0
            sold = 0
            spent = 0
            gain = 0
            if id is not None and name is not None and price1 and remain is not None and sold is not None:
                try:
                    c=conn.cursor()
                    c.execute("INSERT INTO product (product_id, product_name, product_price, product_remain, product_sold, product_spent, product_gain) VALUES (%s, %s, %s, %s, %s, %s, %s)", (id, name, price1, remain, sold, spent, gain))
                    conn.commit()
                    insert()
                    print("Succes")
                except Exception as e:
                    messagebox.showerror("Error","Syntax type error")
            else:
                messagebox.showerror("Error", "Non-existing item")
            clear_all_entrys()
        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()
                for item in selected_item:
                    values = tree.item(item, 'values')
                    values1 = values[0]

                    cursor = conn.cursor()
                    # Начало транзакции
                    cursor.execute("START TRANSACTION;")
                # Выполнение SQL-запроса для удаления строки по определенному идентификатору
                cursor.execute("DELETE FROM product WHERE \"product_id\" = %s", (values1,))
                conn.commit()
                insert()  # Обновить вывод таблицы после добавления строки
                print("The string was successfully deleted from the database")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a string from the database {e}"
                print(error_msg)
                messagebox.showerror("Error when adding a string to the database", error_msg)

        def update_string():
            id_value = int(entry_id.get())
            name = entry_name.get()
            price1 = int(entry_price1.get())
            if not id_value:
                messagebox.showerror("Error", "Please enter the correct values for the product ID")
                return

            try:
                id_value = int(id_value)
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE product SET product_name = %s, product_price = %s WHERE product_id = %s", (name, price1, id_value))
                conn.commit()
                # Обновляем поля после успешного обновления
                clear_all_entrys()
                insert()  # Может потребоваться обновить данные на экране

            except ValueError:
                messagebox.showerror("Error", "The product ID and the number must be integer values")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error updating data {e}"
                print(error_msg)
                messagebox.showerror("Error", error_msg)

            return entry_id, entry_name, entry_price1, \
                   # entry_remain, entry_sold

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id.delete(0, tk.END)
                entry_id.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                entry_name.delete(0, tk.END)
                entry_name.insert(0, item_values[1])
                entry_price1.delete(0, tk.END)
                entry_price1.insert(0, item_values[2])
        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP, padx=6)
        # Change String
        button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_left, bg="#C0C0C0" ,text="Sort", padx=5, pady=1, width=10, height=1, command=sort_by_id)
        button_sort.pack(side=tk.TOP)

        frame_id = tk.Frame(root)
        frame_id.pack(anchor="nw")
        label_id = tk.Label(frame_id, text="Enter id:", width=10, height=1, background="#FFFFFF", anchor='nw')
        label_id.pack(side="left")
        entry_id=tk.Entry(frame_id, width=84)
        entry_id.pack(side="left")

        frame_name = tk.Frame(root)
        frame_name.pack(anchor="nw")
        label_name = tk.Label(frame_name, text="Enter name:", width=10, height=1, background="#FFFFFF", anchor='nw')
        label_name.pack(side="left")
        entry_name=tk.Entry(frame_name, width=84)
        entry_name.pack(side="left")

        frame_price1 = tk.Frame(root)
        frame_price1.pack(anchor="nw")
        label_price1 = tk.Label(frame_price1, text="Enter price:", width=10, height=1, background="#FFFFFF", anchor='nw')
        label_price1.pack(side="left")
        entry_price1 = tk.Entry(frame_price1, width=84)
        entry_price1.pack(side="left")

        frame_sort = tk.Frame(root)
        frame_sort.pack(anchor="nw")
        label_sort = tk.Label(frame_sort, text="Id to find:", width=10, height=2, background="#C0C0C0", anchor='nw')
        label_sort.pack(side="left")
        entry_sort = tk.Entry(frame_sort, width=84)
        entry_sort.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()

def arrival():

    def clear_all_entrys():
        entry_id1.delete(0, tk.END)
        entry_id.delete(0, tk.END)
        entry_date.delete(0, tk.END)
        entry_amount.delete(0, tk.END)
        entry_weight.delete(0, tk.END)
        entry_price.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Arrival")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Arrival id', 'Product id', 'Arrival date', 'Arrival amount', 'Arrival weight', 'Arrival price')
    tree.heading('#0', text='Number')
    tree.heading('Arrival id', text='Arrival id')
    tree.heading('Product id', text='Product id')
    tree.heading('Arrival date', text='Arrival date')
    tree.heading('Arrival amount', text='Arrival amount')
    tree.heading('Arrival weight', text='Arrival weight')
    tree.heading('Arrival price', text='Arrival price')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=60)
    tree.column("Arrival id", width=70, anchor=tk.CENTER)
    tree.column("Product id", width=120, anchor=tk.CENTER)
    tree.column("Arrival date", width=80, anchor=tk.CENTER)
    tree.column("Arrival amount", width=95, anchor=tk.CENTER)
    tree.column("Arrival weight", width=85, anchor=tk.CENTER)
    tree.column("Arrival price", width=85, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM arrival")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM arrival")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM arrival")
        rows = cursor.fetchall()

    # Sorting
    def sort_by_id():
        try:
            selected_product_id = int(entry_sort.get())
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
            'Arrival id', 'Product id', 'Arrival date', 'Arrival amount', 'Arrival weight', 'Arrival price')
            result_tree.heading('#0', text='№')
            result_tree.heading('Arrival id', text='Arrival id')
            result_tree.heading('Product id', text='Product id')
            result_tree.heading('Arrival date', text='Arrival date')
            result_tree.heading('Arrival amount', text='Arrival amount')
            result_tree.heading('Arrival weight', text='Arrival weight')
            result_tree.heading('Arrival price', text='Arrival price')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Arrival id", width=100, anchor=tk.CENTER)
            result_tree.column("Product id", width=220, anchor=tk.CENTER)
            result_tree.column("Arrival date", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival amount", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival weight", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival price", width=100, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM arrival WHERE product_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

            def save_to_excel(selected_product_id):
                cursor.execute("SELECT * FROM arrival WHERE product_id=%s", (selected_product_id,))
                rows = cursor.fetchall()
                wb = Workbook()
                ws = wb.active
                ws.append(
                    ['Arrival id', 'Product id', 'Arrival date', 'Arrival amount', 'Arrival weight', 'Arrival price'])
                for row in rows:
                    ws.append(row)
                desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

                wb.save(save_path)
                print(f"The data has been successfully saved in the file {save_path} on the desktop.")

            # Создание кнопки "Save"
            button_save = tk.Button(result_window, text='Save', width=10, height=1,
                                    command=lambda: save_to_excel(selected_product_id))
            button_save.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        # Sorting
    def sort_by_date():
        try:
            selected_product_id = entry_sort_date.get()
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
                'Arrival id', 'Product id', 'Arrival date', 'Arrival amount', 'Arrival weight', 'Arrival price')
            result_tree.heading('#0', text='№')
            result_tree.heading('Arrival id', text='Arrival id')
            result_tree.heading('Product id', text='Product id')
            result_tree.heading('Arrival date', text='Arrival date')
            result_tree.heading('Arrival amount', text='Arrival amount')
            result_tree.heading('Arrival weight', text='Arrival weight')
            result_tree.heading('Arrival price', text='Arrival price')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Arrival id", width=100, anchor=tk.CENTER)
            result_tree.column("Product id", width=220, anchor=tk.CENTER)
            result_tree.column("Arrival date", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival amount", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival weight", width=100, anchor=tk.CENTER)
            result_tree.column("Arrival price", width=100, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM arrival WHERE arrival_date=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        def save_to_excel(selected_product_id):
            cursor.execute("SELECT * FROM arrival WHERE arrival_date=%s", (selected_product_id,))
            rows = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            ws.append(['Arrival id', 'Product id', 'Arrival date', 'Arrival amount', 'Arrival weight', 'Arrival price'])
            for row in rows:
                ws.append(row)
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            wb.save(save_path)
            print(f"The data has been successfully saved in the file {save_path} on the desktop.")

        # Создание кнопки "Save"
        button_save = tk.Button(result_window, text='Save', width=10, height=1,command=lambda: save_to_excel(selected_product_id))
        button_save.pack()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"arrival_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()
        def add_string():
            id1 = int(entry_id1.get())
            id = int(entry_id.get())
            date = entry_date.get()
            amount = int(entry_amount.get())
            weight = int(entry_weight.get())
            price = int(entry_price.get())

            if id1 is not None and id is not None and date and amount is not None and weight is not None and price is not None:
                try:
                    c=conn.cursor()
                    c.execute("INSERT INTO arrival (arrival_id, product_id, arrival_date, arrival_amount, arrival_weight, arrival_price) VALUES (%s, %s, %s, %s, %s, %s)", ((id1, id, date, amount, weight, price)))

                    if weight==0:
                        # Обновляем произведением arrival_price на arrival_amount
                        spent = amount * price
                        c.execute("UPDATE product SET product_spent = product_spent + %s WHERE product_id = %s",
                                  (spent, id))
                    else:
                        # Обновляем произведением arrival_price на arrival_amount
                        spent = weight * price
                        c.execute("UPDATE product SET product_spent = product_spent + %s WHERE product_id = %s",
                                  (spent, id))

                    # Получаем текущее значение product_remain
                    cursor = conn.cursor()
                    cursor.execute("SELECT product_remain FROM product WHERE product_id = %s", (id,))
                    current_remain = cursor.fetchone()

                    if current_remain:
                        current_remain = current_remain[0]
                        current_remain = int(current_remain)

                        # Обновляем product_remain суммой текущего значения и amount
                        new_remain = current_remain + amount + weight
                        cursor.execute("UPDATE product SET product_remain = %s WHERE product_id = %s", (new_remain, id))

                        conn.commit()
                        insert()
                        print("Success")
                    else:
                        messagebox.showerror("Error", "No current remain for this product")
                except Exception as e:
                    messagebox.showerror("Error","Syntax type error")
            else:
                messagebox.showerror("Error", "Non-existing item")
            clear_all_entrys()
        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()

                for item in selected_item:
                    values = tree.item(item, 'values')
                    values1 = values[0]

                    cursor.execute("START TRANSACTION;")

                    cursor.execute(
                        "SELECT product_id, arrival_amount, arrival_weight, arrival_price FROM arrival WHERE arrival_id = %s",
                        (values1,))
                    result = cursor.fetchone()

                    if result:
                        product_id = result[0]
                        amount = result[1]
                        weight = result[2]
                        spent = result[3] * amount  # calculate the spent for this specific record

                        # Update product_spent to deduct the spent for this record
                        cursor.execute("UPDATE product SET product_spent = product_spent - %s WHERE product_id = %s",
                                       (spent, product_id))

                        # Update product_remain for the corresponding product
                        cursor.execute(
                            "UPDATE product SET product_remain = product_remain - %s - %s WHERE product_id = %s",
                            (amount, weight, product_id))

                        # Delete the record from arrival
                        cursor.execute("DELETE FROM arrival WHERE arrival_id = %s", (values1,))

                conn.commit()
                insert()  # Обновить вывод таблицы после добавления строки
                print("The string was successfully deleted from the database")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a string from the database {e}"
                print(error_msg)
                messagebox.showerror("Error when adding a string to the database", error_msg)

        def update_string():
            id1 = int(entry_id1.get())
            id = int(entry_id.get())
            date = entry_date.get()
            amount = int(entry_amount.get())
            weight = int(entry_weight.get())
            price = int(entry_price.get())

            if not id1:
                messagebox.showerror("Error", "Please enter the correct values for the product ID")
                return

            try:
                cursor = conn.cursor()

                # Получаем текущие данные arrival перед обновлением
                cursor.execute(
                    "SELECT product_id, arrival_amount, arrival_weight, arrival_price FROM arrival WHERE arrival_id = %s",
                    (id1,))
                old_data = cursor.fetchone()

                if old_data:
                    old_product_id = old_data[0]
                    old_amount = old_data[1]
                    old_weight = old_data[2]
                    old_spent = old_data[3] * old_amount

                    # Обновляем product_spent, чтобы отразить изменения в arrival
                    cursor.execute("UPDATE product SET product_spent = product_spent - %s WHERE product_id = %s",
                                   (old_spent, old_product_id))

                    # Уменьшаем product_remain на старые значения arrival перед изменением
                    cursor.execute("UPDATE product SET product_remain = product_remain - %s - %s WHERE product_id = %s",
                                   (old_amount, old_weight, old_product_id))

                    # Обновляем данные в arrival
                    cursor.execute(
                        "UPDATE arrival SET product_id = %s, arrival_date = %s, arrival_amount = %s, arrival_weight = %s, arrival_price = %s WHERE arrival_id = %s",
                        (id, date, amount, weight, price, id1))

                    # Рассчитываем новые значения и обновляем product_spent и product_remain
                    spent = price * amount
                    cursor.execute("UPDATE product SET product_spent = product_spent + %s WHERE product_id = %s",
                                   (spent, id))
                    new_remain = amount + weight
                    cursor.execute("UPDATE product SET product_remain = product_remain + %s WHERE product_id = %s",
                                   (new_remain, id))

                    conn.commit()
                    clear_all_entrys()
                    insert()
                    print("Success")

                else:
                    messagebox.showerror("Error", "Data not found for this arrival ID")

            except ValueError:
                messagebox.showerror("Error", "The product ID and the number must be integer values")

            except Exception as e:
                conn.rollback()
                error_msg = f"Error updating data {e}"
                print(error_msg)
                messagebox.showerror("Error", error_msg)

            return entry_id1, entry_id, entry_date, entry_amount, entry_weight, entry_price

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id1.delete(0, tk.END)
                entry_id1.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                entry_id.delete(0, tk.END)
                entry_id.insert(0, item_values[1])
                entry_date.delete(0, tk.END)
                entry_date.insert(0, item_values[2])
                entry_amount.delete(0, tk.END)
                entry_amount.insert(0, item_values[3])
                entry_weight.delete(0, tk.END)
                entry_weight.insert(0, item_values[4])
                entry_price.delete(0, tk.END)
                entry_price.insert(0, item_values[5])

        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_right, bg="#C0C0C0", text="Sort by date", padx=5, pady=1, width=10, height=1,command=sort_by_date)
        button_sort.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP)
        # Change String
        button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_left, bg="#C0C0C0", text="Sort by id", padx=5, pady=1, width=10, height=1, command=sort_by_id)
        button_sort.pack(side=tk.TOP)

        frame_id1 = tk.Frame(root)
        frame_id1.pack(anchor="nw")
        label_id1 = tk.Label(frame_id1, text="Enter arrival id:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_id1.pack(side="left")
        entry_id1=tk.Entry(frame_id1, width=50)
        entry_id1.pack(side="left")

        frame_id = tk.Frame(root)
        frame_id.pack(anchor="nw")
        label_id = tk.Label(frame_id, text="Enter product id:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_id.pack(side="left")
        entry_id=tk.Entry(frame_id, width=50)
        entry_id.pack(side="left")

        frame_date = tk.Frame(root)
        frame_date.pack(anchor="nw")
        label_date = tk.Label(frame_date, text="Enter date:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_date.pack(side="left")
        entry_date = tk.Entry(frame_date, width=50)
        entry_date.pack(side="left")

        frame_amount = tk.Frame(root)
        frame_amount.pack(anchor="nw")
        label_amount = tk.Label(frame_amount, text="Enter amount:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_amount.pack(side="left")
        entry_amount = tk.Entry(frame_amount, width=50)
        entry_amount.pack(side="left")

        frame_weight = tk.Frame(root)
        frame_weight.pack(anchor="nw")
        label_weight = tk.Label(frame_weight, text="Enter weight:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_weight.pack(side="left")
        entry_weight = tk.Entry(frame_weight, width=50)
        entry_weight.pack(side="left")

        frame_price = tk.Frame(root)
        frame_price.pack(anchor="nw")
        label_price = tk.Label(frame_price, text="Enter price:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_price.pack(side="left")
        entry_price = tk.Entry(frame_price, width=50)
        entry_price.pack(side="left")

        frame_sort = tk.Frame(root)
        frame_sort.pack(anchor="nw")
        label_sort = tk.Label(frame_sort, text="Product to find:", width=13, height=1, background="#C0C0C0", anchor='nw')
        label_sort.pack(side="left")
        entry_sort = tk.Entry(frame_sort, width=50)
        entry_sort.pack(side="left")

        frame_sort_date = tk.Frame(root)
        frame_sort_date.pack(anchor="nw")
        label_sort_date = tk.Label(frame_sort_date, text="Date to find:", width=13, height=1, background="#C0C0C0",anchor='nw')
        label_sort_date.pack(side="left")
        entry_sort_date = tk.Entry(frame_sort_date, width=50)
        entry_sort_date.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()

def client():
    def clear_all_entrys():
        entry_id2.delete(0, tk.END)
        entry_name1.delete(0, tk.END)
        entry_email.delete(0, tk.END)
        entry_address.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Clients")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Client id','Client name','Client email','Client address')
    tree.heading('#0', text='Number')
    tree.heading('Client id', text='Client id')
    tree.heading('Client name', text='Client name')
    tree.heading('Client email', text='Client email')
    tree.heading('Client address', text='Client address')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=60)
    tree.column("Client id", width=70, anchor=tk.CENTER)
    tree.column("Client name", width=220, anchor=tk.CENTER)
    tree.column("Client email", width=80, anchor=tk.CENTER)
    tree.column("Client address", width=95, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM client")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM client")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM client")
        rows = cursor.fetchall()

    # Sorting
    def sort_by_id():
        try:
            selected_product_id = entry_sort.get()
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
                    'Client id','Client name','Client email','Client address')
            result_tree.heading('#0', text='№')
            result_tree.heading('Client id', text='Client id')
            result_tree.heading('Client name', text='Client name')
            result_tree.heading('Client email', text='Client email')
            result_tree.heading('Client address', text='Client address')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Client id", width=100, anchor=tk.CENTER)
            result_tree.column("Client name", width=220, anchor=tk.CENTER)
            result_tree.column("Client email", width=100, anchor=tk.CENTER)
            result_tree.column("Client address", width=100, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM client WHERE client_name=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        def save_to_excel(selected_product_id):
            cursor.execute("SELECT * FROM client WHERE client_name=%s", (selected_product_id,))
            rows = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            ws.append(
                ['Client id','Client name','Client email','Client address'])
            for row in rows:
                ws.append(row)
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            wb.save(save_path)
            print(f"The data has been successfully saved in the file {save_path} on the desktop.")

        # Создание кнопки "Save"
        button_save = tk.Button(result_window, text='Save', width=10, height=1,
                                    command=lambda: save_to_excel(selected_product_id))
        button_save.pack()



    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"client_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()
        def add_string():
            id2 = entry_id2.get()
            name1 = entry_name1.get()
            email = entry_email.get()
            address = entry_address.get()
            if id2 is not None and name1 is not None and email and address is not None :
                try:
                    c=conn.cursor()
                    c.execute("INSERT INTO client (client_id, client_name, client_email, client_address) VALUES (%s, %s, %s, %s)", ((id2, name1, email, address)))
                    conn.commit()
                    insert()
                    print("Succes")
                except Exception as e:
                    messagebox.showerror("Error","Syntax type error")
            else:
                messagebox.showerror("Error", "Non-existing item")
            clear_all_entrys()
        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()

                for item in selected_item:
                    values = tree.item(item, 'values')
                    values1 = values[0]

                    cursor = conn.cursor()
                    # Начало транзакции
                    cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для удаления строки по определенному идентификатору
                cursor.execute("DELETE FROM client WHERE \"client_id\" = %s", (values1,))

                conn.commit()
                insert()  # Обновить вывод таблицы после добавления строки
                print("The string was successfully deleted from the database")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a string from the database {e}"
                print(error_msg)
                messagebox.showerror("Error when adding a string to the database", error_msg)

        def update_string():
            id2 = int(entry_id2.get())
            name1 = entry_name1.get()
            email = entry_email.get()
            address = entry_address.get()
            if not id2:
                messagebox.showerror("Error", "Please enter the correct values for the product ID")
                return

            try:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE client SET client_name = %s, client_email = %s, client_address = %s WHERE client_id = %s", (name1, email, address, id2))
                conn.commit()
                # Обновляем поля после успешного обновления
                clear_all_entrys()
                insert()  # Может потребоваться обновить данные на экране

            except ValueError:
                messagebox.showerror("Error", "The product ID and the number must be integer values")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error updating data {e}"
                print(error_msg)
                messagebox.showerror("Error", error_msg)

            return name1, email, address, id2

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id2.delete(0, tk.END)
                entry_id2.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                entry_name1.delete(0, tk.END)
                entry_name1.insert(0, item_values[1])
                entry_email.delete(0, tk.END)
                entry_email.insert(0, item_values[2])
                entry_address.delete(0, tk.END)
                entry_address.insert(0, item_values[3])
        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP)
        # Change String
        button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_left,bg="#C0C0C0", text="Sort by name", padx=5, pady=1, width=10, height=1, command=sort_by_id)
        button_sort.pack(side=tk.TOP)

        frame_id2 = tk.Frame(root)
        frame_id2.pack(anchor="nw")
        label_id2 = tk.Label(frame_id2, text="Enter id:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_id2.pack(side="left")
        entry_id2=tk.Entry(frame_id2, width=38)
        entry_id2.pack(side="left")

        frame_name1 = tk.Frame(root)
        frame_name1.pack(anchor="nw")
        label_name1 = tk.Label(frame_name1, text="Enter name:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_name1.pack(side="left")
        entry_name1=tk.Entry(frame_name1, width=38)
        entry_name1.pack(side="left")

        frame_email = tk.Frame(root)
        frame_email.pack(anchor="nw")
        label_email = tk.Label(frame_email, text="Enter email:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_email.pack(side="left")
        entry_email = tk.Entry(frame_email, width=38)
        entry_email.pack(side="left")

        frame_address = tk.Frame(root)
        frame_address.pack(anchor="nw")
        label_address = tk.Label(frame_address, text="Enter address:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_address.pack(side="left")
        entry_address = tk.Entry(frame_address, width=38)
        entry_address.pack(side="left")

        frame_sort = tk.Frame(root)
        frame_sort.pack(anchor="nw")
        label_sort = tk.Label(frame_sort, text="Name to find:", width=14, height=1, background="#C0C0C0",anchor='nw')
        label_sort.pack(side="left")
        entry_sort = tk.Entry(frame_sort, width=38)
        entry_sort.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()

def order():

    def clear_all_entrys():
        entry_id3.delete(0, tk.END)
        entry_amount1.delete(0, tk.END)
        entry_id2.delete(0, tk.END)
        entry_id.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Orders")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Order id','Order amount','Client id','Product id')
    tree.heading('#0', text='Number')
    tree.heading('Order id', text='Order id')
    tree.heading('Order amount', text='Order amount')
    tree.heading('Client id', text='Client id')
    tree.heading('Product id', text='Product id')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=60)
    tree.column("Order id", width=70, anchor=tk.CENTER)
    tree.column("Order amount", width=90, anchor=tk.CENTER)
    tree.column("Client id", width=75, anchor=tk.CENTER)
    tree.column("Product id", width=85, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM \"order\"")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM \"order\"")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM \"order\"")
        rows = cursor.fetchall()

    # Sorting
    def sort_by_id():
        try:
            selected_product_id = int(entry_sort.get())
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
                    'Order id','Order amount','Client id','Product id')
            result_tree.heading('#0', text='№')
            result_tree.heading('Order id', text='Order id')
            result_tree.heading('Order amount', text='Order amount')
            result_tree.heading('Client id', text='Client id')
            result_tree.heading('Product id', text='Product id')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Order id", width=100, anchor=tk.CENTER)
            result_tree.column("Order amount", width=100, anchor=tk.CENTER)
            result_tree.column("Client id", width=100, anchor=tk.CENTER)
            result_tree.column("Product id", width=100, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM \"order\" WHERE client_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        def save_to_excel(selected_product_id):
            cursor.execute("SELECT * FROM \"order\" WHERE client_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            ws.append(
                ['Order id','Order amount','Client id','Product id'])
            for row in rows:
                ws.append(row)
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            wb.save(save_path)
            print(f"The data has been successfully saved in the file {save_path} on the desktop.")

        # Создание кнопки "Save"
        button_save = tk.Button(result_window, text='Save', width=10, height=1,
                                    command=lambda: save_to_excel(selected_product_id))
        button_save.pack()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"order_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()
        def add_string():
            id3 = int(entry_id3.get())
            amount1 = int(entry_amount1.get())
            id2 = int(entry_id2.get())
            id = int(entry_id.get())
            if id3 is not None and amount1 is not None and id2 and id is not None :
                try:
                    c=conn.cursor()
                    c.execute("INSERT INTO \"order\" (order_id, order_amount, client_id, product_id) VALUES (%s, %s, %s, %s)", ((id3, amount1, id2, id)))
                    conn.commit()
                    insert()
                    print("Succes")
                except Exception as e:
                    messagebox.showerror("Error","Syntax type error")
            else:
                messagebox.showerror("Error", "Non-existing item")
            clear_all_entrys()
        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()

                for item in selected_item:
                    values = tree.item(item, 'values')
                    values1 = values[0]

                    cursor = conn.cursor()
                    # Начало транзакции
                    cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для удаления строки по определенному идентификатору
                cursor.execute("DELETE FROM \"order\" WHERE \"order_id\" = %s", (values1,))

                conn.commit()
                insert()  # Обновить вывод таблицы после добавления строки
                print("The string was successfully deleted from the database")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a string from the database {e}"
                print(error_msg)
                messagebox.showerror("Error when adding a string to the database", error_msg)

        def update_string():
            id3 = int(entry_id3.get())
            amount1 = entry_amount1.get()
            id2 = int(entry_id2.get())
            id = int(entry_id.get())
            if not id3:
                messagebox.showerror("Error", "Please enter the correct values for the order ID")
                return

            try:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE \"order\" SET order_amount = %s, client_id = %s, product_id = %s WHERE order_id = %s", (amount1, id2, id, id3))
                conn.commit()
                # Обновляем поля после успешного обновления
                clear_all_entrys()
                insert()  # Может потребоваться обновить данные на экране

            except ValueError:
                messagebox.showerror("Error", "The product ID and the number must be integer values")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error updating data {e}"
                print(error_msg)
                messagebox.showerror("Error", error_msg)

            return amount1, id2, id, id3

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id3.delete(0, tk.END)
                entry_id3.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                entry_amount1.delete(0, tk.END)
                entry_amount1.insert(0, item_values[1])
                entry_id2.delete(0, tk.END)
                entry_id2.insert(0, item_values[2])
                entry_id.delete(0, tk.END)
                entry_id.insert(0, item_values[3])
        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP)
        # Change String
        button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_left,bg="#C0C0C0", text="Sort by id", padx=5, pady=1, width=10, height=1, command=sort_by_id)
        button_sort.pack(side=tk.TOP)

        frame_id3 = tk.Frame(root)
        frame_id3.pack(anchor="nw")
        label_id3 = tk.Label(frame_id3, text="Enter order id:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_id3.pack(side="left")
        entry_id3=tk.Entry(frame_id3, width=15)
        entry_id3.pack(side="left")

        frame_amount1 = tk.Frame(root)
        frame_amount1.pack(anchor="nw")
        label_amount1 = tk.Label(frame_amount1, text="Enter amount:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_amount1.pack(side="left")
        entry_amount1=tk.Entry(frame_amount1, width=15)
        entry_amount1.pack(side="left")

        frame_id2 = tk.Frame(root)
        frame_id2.pack(anchor="nw")
        label_id2 = tk.Label(frame_id2, text="Enter client id:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_id2.pack(side="left")
        entry_id2 = tk.Entry(frame_id2, width=15)
        entry_id2.pack(side="left")

        frame_id = tk.Frame(root)
        frame_id.pack(anchor="nw")
        label_id = tk.Label(frame_id, text="Enter product id:", width=13, height=1, background="#FFFFFF", anchor='nw')
        label_id.pack(side="left")
        entry_id = tk.Entry(frame_id, width=15)
        entry_id.pack(side="left")

        frame_sort = tk.Frame(root)
        frame_sort.pack(anchor="nw")
        label_sort = tk.Label(frame_sort, text="Clent id to find:", width=13, height=1, background="#C0C0C0",
                              anchor='nw')
        label_sort.pack(side="left")
        entry_sort = tk.Entry(frame_sort, width=15)
        entry_sort.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()

def payment():

    def clear_all_entrys():
        entry_id4.delete(0, tk.END)
        entry_id3.delete(0, tk.END)
        entry_date1.delete(0, tk.END)
        # entry_price2.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Payments")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Payment id','Order id','Payment date','Payment price','Product id')
    tree.heading('#0', text='Number')
    tree.heading('Payment id', text='Payment id')
    tree.heading('Order id', text='Order id')
    tree.heading('Payment date', text='Payment date')
    tree.heading('Payment price', text='Payment price')
    tree.heading('Product id', text='Product id')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=60)
    tree.column("Payment id", width=70, anchor=tk.CENTER)
    tree.column("Order id", width=70, anchor=tk.CENTER)
    tree.column("Payment date", width=80, anchor=tk.CENTER)
    tree.column("Payment price", width=95, anchor=tk.CENTER)
    tree.column("Product id", width=70, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM payment")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM payment")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM payment")
        rows = cursor.fetchall()

    # Sorting by id
    def sort_by_id():
        try:
            selected_product_id = int(entry_sort.get())
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
                    'Payment id','Order id','Payment date','Payment price','Product id')
            result_tree.heading('#0', text='№')
            result_tree.heading('Payment id', text='Payment id')
            result_tree.heading('Order id', text='Order id')
            result_tree.heading('Payment date', text='Payment date')
            result_tree.heading('Payment price', text='Payment price')
            result_tree.heading('Product id', text='Product id')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Payment id", width=100, anchor=tk.CENTER)
            result_tree.column("Order id", width=100, anchor=tk.CENTER)
            result_tree.column("Payment date", width=100, anchor=tk.CENTER)
            result_tree.column("Payment price", width=100, anchor=tk.CENTER)
            result_tree.column("Product id", width=80, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM payment WHERE order_id=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

            def save_to_excel(selected_product_id):
                cursor.execute("SELECT * FROM payment WHERE order_id=%s", (selected_product_id,))
                rows = cursor.fetchall()
                wb = Workbook()
                ws = wb.active
                ws.append(
                    ['Payment id', 'Order id', 'Payment date', 'Payment price'])
                for row in rows:
                    ws.append(row)
                desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

                wb.save(save_path)
                print(f"The data has been successfully saved in the file {save_path} on the desktop.")

            # Создание кнопки "Save"
            button_save = tk.Button(result_window, text='Save', width=10, height=1,
                                    command=lambda: save_to_excel(selected_product_id))
            button_save.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

    # Sorting by date
    def sort_by_date():
        try:
            selected_product_id = str(entry_sort_date.get())
            # Создаем новое окно для отображения сортированных данных
            result_window = tk.Toplevel(root)
            result_window.title("Sorted by product id")
            result_window.resizable(False, False)
            result_window.configure(bg='#FFFFFF')

            result_tree = ttk.Treeview(result_window)
            result_tree["columns"] = (
                'Payment id', 'Order id', 'Payment date', 'Payment price', 'Product id')
            result_tree.heading('#0', text='№')
            result_tree.heading('Payment id', text='Payment id')
            result_tree.heading('Order id', text='Order id')
            result_tree.heading('Payment date', text='Payment date')
            result_tree.heading('Payment price', text='Payment price')
            result_tree.heading('Product id', text='Product id')

            # Устанавливаем размеры столбцов для таблицы результатов
            result_tree.column("#0", width=40)
            result_tree.column("Payment id", width=100, anchor=tk.CENTER)
            result_tree.column("Order id", width=100, anchor=tk.CENTER)
            result_tree.column("Payment date", width=100, anchor=tk.CENTER)
            result_tree.column("Payment price", width=100, anchor=tk.CENTER)
            result_tree.column("Product id", width=80, anchor=tk.CENTER)

            cursor.execute("SELECT * FROM payment WHERE payment_date=%s", (selected_product_id,))
            rows = cursor.fetchall()
            for i, row in enumerate(rows):
                result_tree.insert("", "end", text=str(i), values=row)
            result_tree.pack()

        except ValueError as e:
            print(f"Error: {e}. Enter correct product id")

        def save_to_excel(selected_product_id):
            cursor.execute("SELECT * FROM payment WHERE payment_date=%s", (selected_product_id,))
            rows = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            ws.append(
                ['Payment id','Order id','Payment date','Payment price'])
            for row in rows:
                ws.append(row)
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"product_data_{timestamp}.xlsx")

            wb.save(save_path)
            print(f"The data has been successfully saved in the file {save_path} on the desktop.")

        # Создание кнопки "Save"
        button_save = tk.Button(result_window, text='Save', width=10, height=1,command=lambda: save_to_excel(selected_product_id))
        button_save.pack()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"payment_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()
        def add_string():
            id4 = int(entry_id4.get())
            id3 = int(entry_id3.get())
            date1 = entry_date1.get()
            if id4 is not None and id3 is not None and date1:
                try:
                    c = conn.cursor()
                    # Получение product_id и product_price для соответствующего order_id
                    c.execute(
                        "SELECT product.product_id, product.product_price FROM product JOIN \"order\" ON product.product_id = \"order\".product_id WHERE \"order\".order_id = %s",
                        (id3,))
                    row = c.fetchone()

                    if row:
                        product_id = row[0]
                        product_price = row[1]

                        # Получение order_amount для данного order_id
                        c.execute("SELECT order_amount FROM \"order\" WHERE order_id = %s", (id3,))
                        row2 = c.fetchone()

                        if row2:
                            order_amount = row2[0]
                            payment_price = order_amount * product_price

                            # Вставка данных о платеже в таблицу payment
                            c.execute(
                                "INSERT INTO payment (payment_id, order_id, payment_date, payment_price, product_id) VALUES (%s, %s, %s, %s, %s)",
                                (id4, id3, date1, payment_price, product_id))
                            conn.commit()

                            # Обновление product_gain в таблице product
                            c.execute("UPDATE product SET product_gain = product_gain + %s WHERE product_id = %s",
                                      (payment_price, product_id))
                            conn.commit()

                            insert()
                            print("Success")
                        else:
                            messagebox.showerror("Error", "Incorrect order_id")
                    else:
                        messagebox.showerror("Error", "Product data not found for the given order_id")
                except Exception as e:
                    messagebox.showerror("Error", "An error occurred: {}".format(str(e)))
            else:
                messagebox.showerror("Error", "Non-existing item")
            clear_all_entrys()

        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()

                for item in selected_item:
                    values = tree.item(item, 'values')
                    payment_id = values[0]

                    cursor.execute("START TRANSACTION;")

                    # Retrieve the necessary information from the payment table before deletion
                    cursor.execute("SELECT product_id, payment_price FROM payment WHERE payment_id = %s", (payment_id,))
                    row = cursor.fetchone()

                    if row:
                        product_id = row[0]
                        payment_price = row[1]

                        # Delete the selected row from the payment table
                        cursor.execute("DELETE FROM payment WHERE payment_id = %s", (payment_id,))

                        # Update product_gain in the product table
                        cursor.execute("UPDATE product SET product_gain = product_gain - %s WHERE product_id = %s",
                                       (payment_price, product_id))

                        conn.commit()
                        insert()  # Update the displayed table after deletion
                        print("The row was successfully deleted from the database")

                    else:
                        print("Payment data not found for the given payment_id")

            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a row from the database: {e}"
                print(error_msg)
                messagebox.showerror("Error when deleting a row from the database", error_msg)

        def update_string():
            id4 = int(entry_id4.get())
            id3 = int(entry_id3.get())
            date1 = entry_date1.get()
            if not id4:
                messagebox.showerror("Error", "Please enter the correct values for the payment ID")
                return

            try:
                cursor = conn.cursor()

                # Получение product_id и product_price для соответствующего order_id
                cursor.execute(
                    "SELECT product.product_id, product.product_price FROM product JOIN \"order\" ON product.product_id = \"order\".product_id WHERE \"order\".order_id = %s",
                    (id3,))
                row = cursor.fetchone()

                if row:
                    product_id = row[0]
                    product_price = row[1]

                    # Получение order_amount для данного order_id
                    cursor.execute("SELECT order_amount FROM \"order\" WHERE order_id = %s", (id3,))
                    row2 = cursor.fetchone()

                    if row2:
                        order_amount = row2[0]
                        payment_price = order_amount * product_price

                        # Обновление данных в таблице payment
                        cursor.execute(
                            "UPDATE payment SET order_id = %s, payment_date = %s, payment_price = %s, product_id = %s WHERE payment_id = %s",
                            (id3, date1, payment_price, product_id, id4))
                        conn.commit()

                        insert()
                        print("Success")
                    else:
                        messagebox.showerror("Error", "Incorrect order_id")
                else:
                    messagebox.showerror("Error", "Product data not found for the given order_id")

            except ValueError:
                messagebox.showerror("Error", "The order ID and the payment ID must be integer values")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error updating data: {e}"
                print(error_msg)
                messagebox.showerror("Error", error_msg)

            clear_all_entrys()
            return id3, date1, payment_price, product_id, id4

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id4.delete(0, tk.END)
                entry_id4.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                entry_id3.delete(0, tk.END)
                entry_id3.insert(0, item_values[1])
                entry_date1.delete(0, tk.END)
                entry_date1.insert(0, item_values[2])
                # entry_price2.delete(0, tk.END)
                # entry_price2.insert(0, item_values[3])
        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_right, bg="#C0C0C0", text="Sort by date", padx=5, pady=1, width=10, height=1,
                                command=sort_by_date)
        button_sort.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP)
        # Change String
        button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        button_sort = tk.Button(button_frame_left, bg="#C0C0C0", text="Sort by id", padx=5, pady=1, width=10, height=1, command=sort_by_id)
        button_sort.pack(side=tk.TOP)

        frame_id4 = tk.Frame(root)
        frame_id4.pack(anchor="nw")
        label_id4 = tk.Label(frame_id4, text="Enter payment id:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_id4.pack(side="left")
        entry_id4 = tk.Entry(frame_id4, width=24)
        entry_id4.pack(side="left")

        frame_id3 = tk.Frame(root)
        frame_id3.pack(anchor="nw")
        label_id3 = tk.Label(frame_id3, text="Enter order id:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_id3.pack(side="left")
        entry_id3=tk.Entry(frame_id3, width=24)
        entry_id3.pack(side="left")

        frame_date1 = tk.Frame(root)
        frame_date1.pack(anchor="nw")
        label_date1 = tk.Label(frame_date1, text="Enter date:", width=14, height=1, background="#FFFFFF", anchor='nw')
        label_date1.pack(side="left")
        entry_date1 = tk.Entry(frame_date1, width=24)
        entry_date1.pack(side="left")

        frame_sort = tk.Frame(root)
        frame_sort.pack(anchor="nw")
        label_sort = tk.Label(frame_sort, text="Order id to find:", width=14, height=1, background="#C0C0C0",anchor='nw')
        label_sort.pack(side="left")
        entry_sort = tk.Entry(frame_sort, width=24)
        entry_sort.pack(side="left")

        frame_sort_date = tk.Frame(root)
        frame_sort_date.pack(anchor="nw")
        label_sort_date = tk.Label(frame_sort_date, text="Order date to find:", width=14, height=1, background="#C0C0C0",anchor='nw')
        label_sort_date.pack(side="left")
        entry_sort_date = tk.Entry(frame_sort_date, width=24)
        entry_sort_date.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()

def profit():

    def clear_all_entrys():
        entry_id.delete(0, tk.END)
        # entry_profit.delete(0, tk.END)

    # Создаем графический интерфейс
    root = tk.Tk()
    root.title("Profit")

    root.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    root.configure(bg='#FFFFFF')  # Задание цвета фона для дочернего окна

    # Создание таблицы Отдел
    tree = ttk.Treeview(root)
    tree["columns"] = ('Product id','Profit')
    tree.heading('#0', text='Number')
    tree.heading('Product id', text='Product id')
    tree.heading('Profit', text='Profit')

    # Устанавливаем размеры столбцов
    tree.column("#0", width=130)
    tree.column("Product id", width=130, anchor=tk.CENTER)
    tree.column("Profit", width=130, anchor=tk.CENTER)

    tree.pack()

    # Предварительная очистка таблицы
    tree.delete(*tree.get_children())
    cursor.execute("SELECT * FROM profit")
    rows = cursor.fetchall()

    # insert для добавления данных в таблицу
    def insert():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM profit")
        rows = cursor.fetchall()
        # Вывод данных в таблицу
        for i, row in enumerate(rows):
            tree.insert("", "end", text=str(i), values=row)

    def clear():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT * FROM profit")
        rows = cursor.fetchall()

    def save_to_excel():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([column[0] for column in cursor.description])  # Заголовки столбцов

            for row_data in rows:
                ws.append(row_data)

            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            save_path = os.path.join(desktop_path, f"profit_data_{timestamp}.xlsx")

            if save_path:
                wb.save(save_path)
                print("The data has been successfully saved to an Excel file")
        except Exception as e:
            print(f"Error when saving data in Excel {e}")

    try:
        insert()

        def add_string():
            id = int(entry_id.get())
            if id is not None:
                try:
                    c = conn.cursor()

                    # Получение суммы значений product_spent и product_gain для данного product_id
                    c.execute("SELECT SUM(product_spent + product_gain) FROM product WHERE product_id = %s", (id,))
                    total_profit = c.fetchone()[0] if c.rowcount else 0

                    # Вставка данных в таблицу profit
                    c.execute("INSERT INTO profit (product_id, profit) VALUES (%s, %s)", (id, total_profit))
                    conn.commit()
                    insert()
                    print("Success")
                except Exception as e:
                    messagebox.showerror("Error", "Non-existing item")
            else:
                messagebox.showerror("Error", "Syntax type error")

            clear_all_entrys()

        def on_enter(event):
            # Поместите здесь ваше действие
            add_string()
        # Привязываем клавишу Enter к функции on_enter
        root.bind('<Return>', on_enter)

        #Delete choosen string
        def delete_string():
            try:
                selected_item = tree.selection()
                cursor = conn.cursor()

                for item in selected_item:
                    values = tree.item(item, 'values')
                    values1 = values[0]

                    cursor = conn.cursor()
                    # Начало транзакции
                    cursor.execute("START TRANSACTION;")

                # Выполнение SQL-запроса для удаления строки по определенному идентификатору
                cursor.execute("DELETE FROM profit WHERE \"product_id\" = %s", (values1,))

                conn.commit()
                insert()  # Обновить вывод таблицы после добавления строки
                print("The string was successfully deleted from the database")
            except Exception as e:
                conn.rollback()
                error_msg = f"Error deleting a string from the database {e}"
                print(error_msg)
                messagebox.showerror("Error when adding a string to the database", error_msg)

        def on_double_click(event):
            selected_item = tree.selection()[0]  # Получаем ID выбранной строки
            item_values = tree.item(selected_item, "values")  # Получаем значения выбранной строки
            if item_values:
                entry_id.delete(0, tk.END)
                entry_id.insert(0, item_values[0])  # Пример: ID отдела в первое поле
                # entry_profit.delete(0, tk.END)
                # entry_profit.insert(0, item_values[1])

        tree.bind("<Double-1>", on_double_click)

        #Delete confirmation
        def confirm_action():
            result = messagebox.askyesno("Delete string", "Are you sure you want to delete the string?")
            if result:
                delete_string()
            else:
                print("Action canceled.")

        button_frame_right = tk.Frame(root, background="#FFFFFF")
        button_frame_right.pack(side='right')
        # Import page
        button_import = tk.Button(button_frame_right, text="Import page", padx=5, pady=1, width=10, height=1, command=insert)
        button_import.pack(side=tk.TOP, padx=6)
        # Clear page
        button_clear = tk.Button(button_frame_right, text="Clear page", padx=5, pady=1, width=10, height=1, command = clear)
        button_clear.pack(side=tk.TOP)
        # Save to excel
        button_excel = tk.Button(button_frame_right, text="Save to excel", padx=5, pady=1, width=10, height=1, command=save_to_excel)
        button_excel.pack(side=tk.TOP)

        button_frame_left = tk.Frame(root, background="#FFFFFF")
        button_frame_left.pack(side='right')
        # Add String
        button_add = tk.Button(button_frame_left, text="Add String", padx=5, pady=1, width=10, height=1, command=add_string)
        button_add.pack(side=tk.TOP)
        # Change String
        # button_change = tk.Button(button_frame_left, text="Change String", padx=5, pady=1, width=10, height=1, command=update_string)
        # button_change.pack(side=tk.TOP)
        # Delete String
        button_delete = tk.Button(button_frame_left, text="Delete String", padx=5, pady=1, width=10, height=1, command= confirm_action)
        button_delete.pack(side=tk.TOP)

        frame_id = tk.Frame(root)
        frame_id.pack(anchor="nw")
        label_id = tk.Label(frame_id, text="Enter product id:", width=13, height=2, background="#FFFFFF", anchor='nw')
        label_id.pack(side="left")
        entry_id = tk.Entry(frame_id, width=16, background="#FFFFFF")
        entry_id.pack(side="left")

        # frame_profit = tk.Frame(root)
        # frame_profit.pack(side="left")
        # label_profit = tk.Label(frame_profit, text="Enter profit:", width=13, height=2, background="#FFFFFF", anchor='nw')
        # label_profit.pack(side="left")
        # entry_profit = tk.Entry(frame_profit, width=16, background="#FFFFFF")
        # entry_profit.pack(side="left")

    except Exception as e:
        print(f"Error loading data {e}")

    # Не трогать
    # Закрытие соединения с базой данных
    root.mainloop()
def documentation ():
    boom1 = tk.Tk()
    boom1.title("Documentation")
    boom1.geometry("700x450")
    boom1.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    boom1.configure(bg='#FFFFFF')  # Задание цвета фона окна в формате #RRGGBB

    label = tk.Label(boom1, text="", font=('Arial', 9, 'normal'), background="#FFFFFF", foreground="white")
    label.pack(pady=5)

    # Создание поля  текстом
    text = tk.Text(boom1)
    text.pack()
    text.tag_configure("custom_font", font=("Arial", 10, "bold"))
    text.insert(tk.END, "# Документация по использованию приложения с базой данных\n\n"
        "## 1. Введение\n"
        "Данная документация предоставляет подробные инструкции по использованию\n"
        "приложения Bratishki Corporation, которое представляет собой\n"
        "информационную систему управления продажами. В этом документе содержится\n"
        "информация о начальной настройке, выполнении основных операций и\n"
        "рекомендациях по эффективному использованию приложения.\n\n"
        "## 2. Начало работы\n\n"
        "### 2.1. Регистрация и вход в приложение\n"
        "Для начала работы с приложением необходимо выполнить следующие шаги:\n\n"
        "1. Регистрация нового пользователя:\n"
        "   - Запустите приложение и нажмите кнопку Registration на главном экране.\n"
        "   - Введите необходимую информацию, включая имя и пароль.\n"
        "   - Подтвердите регистрацию, следуя инструкциям на экране.\n"
        "2. Вход в приложение:\n"
        "   - На главном экране нажмите кнопку Login.\n"
        "   - Введите ваше имя пользователя и пароль, затем нажмите кнопку Login\n"
        "     для доступа к функционалу приложения.\n\n"
        "### 2.2. Подключение к базе данных\n"
        "Приложение автоматически подключается к базе данных при успешном\n"
        "входе в систему. В случае возникновения проблем с подключением\n"
        "убедитесь, что у вас есть стабильное интернет-соединение. Если проблема\n"
        "не решается, обратитесь к системному администратору для получения помощи.\n\n"
        "## 3. Основные операции\n\n"
        "### 3.1. Просмотр данных\n"
        "Для доступа к различным разделам данных используйте основное меню\n"
        "приложения. Основное меню включает следующие разделы:\n\n"
        "- Product - управление данными о продуктах.\n"
        "- Arrival - управление данными о поставках.\n"
        "- Client - данные клиентов.\n"
        "- Order - данные о заказах.\n"
        "- Payment - данные об оплате заказов.\n"
        "- Profit - расчетная конечная окупаемость продукта.\n\n"
        "Для просмотра конкретных записей выберите соответствующий раздел.\n"
        "Например, чтобы просмотреть данные о продуктах, \n"
        "выберите раздел Product.\n\n"
        "### 3.2. Добавление данных\n"
        "Для добавления новых записей в базу данных выполните следующие шаги:\n\n"
        "1. Перейдите в соответствующий раздел (Например Product).\n"
        "2. Заполните необходимые поля в форме (ID Продукта, Имя, Цену за единицу продукта).\n"
        "В качестве одной единицы принимается либо реальная единица,\n"
        "либо 1 кг, если продукт на развес\n"
        "3. Нажмите кнопку Add String для сохранения новой записи в базу данных.\n\n"
        "### 3.3. Редактирование и удаление данных\n"
        "Для редактирования или удаления существующих записей выполните следующие шаги:\n\n"
        "- Редактирование записи:\n"
        "  - Выберите запись, которую необходимо изменить.\n"
        "  - Внесите изменения в соответствующие поля.\n"
        "  - Нажмите кнопку Change для сохранения изменений.\n\n"
        "- Удаление записи:\n"
        "  - Выберите запись, которую необходимо удалить.\n"
        "  - Подтвердите удаление, следуя инструкциям на экране.\n\n"
        "### 3.4. Сортировка и сохранение в виде excel таблицы\n"
        "По соотв. кнопкам Sort by ... и Save to excel, а также Save внутри \n"
        "сортировочного окна предоставлена возможность удобного просмотра \n"
        "данных из таблиц и поиска необходимых значений. \n\n"
        
        "## 4 Безопасность и конфиденциальность\n"
        "- Редактирование неизменяемых данных, таких как расчет\n"
        "конечной прибыли невозможно, эти данные заполняются автоматически.\n"
        "- Распространение информации изнутри базы данных запрещено\n"
        "и преследуется по закону.\n"
        "- Внимательно следите за тем,чтобы возможность управления\n"
        "базой данных не переходила третьим лицам.\n",
        "custom_font")

    boom1.mainloop()

def data():
    def close():
        boom.destroy()

    boom = tk.Tk()
    boom.title("Database")
    boom.geometry("500x750")
    boom.resizable(False, False)  # Это запретит изменение размеров окна по ширине и по высоте
    boom.configure(bg='#0a0a0a')  # Задание цвета фона окна в формате #RRGGBB

    label = tk.Label(boom, text="", font=('Arial', 9, 'normal'), background="#0a0a0a", foreground="white")
    label.pack(pady=0)
    label1 = tk.Label(boom, text="Bratishki Corporation", font=('Arial', 12, 'normal'), background="#0a0a0a", foreground="gray")
    label1.pack(pady=0)
    colon_l = tk.Frame(boom, background="#0a0a0a", padx=20, pady=1)
    colon_l.pack()
    # Создание кнопок
    Product = tk.Button(colon_l, text="Product", command=product, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Product.pack(pady=6)
    Arrival = tk.Button(colon_l, text="Arrival", command=arrival, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Arrival.pack(pady=6)
    Client = tk.Button(colon_l, text="Client", command=client, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Client.pack(pady=6)
    colon_r = tk.Frame(boom, background="#0a0a0a", padx=20, pady=1)
    colon_r.pack()
    Order = tk.Button(colon_r, text="Order", command=order, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Order.pack(pady=6)
    Payment = tk.Button(colon_r, text="Payment", command=payment, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Payment.pack(pady=6)
    Profit = tk.Button(colon_r, text="Profit", command=profit, width=20, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Profit.pack(pady=6)
    Exit_frame = tk.Frame(boom, background="#0a0a0a")
    Exit_frame.pack(side='bottom', pady=10)
    Exit = tk.Button(Exit_frame, text="Exit", command=close, width=10, height=1, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Exit.pack()

def main_menu():

    def close():
        menu.destroy()
    menu = tk.Tk()
    menu.title("Menu")
    menu.geometry("400x380")
    menu.resizable(False, False)
    menu.configure(bg="#0a0a0a")
    label = tk.Label(menu, text="        Bratishki Corporation️", font=('Arial', 14, 'normal'), background="#0a0a0a", foreground="white")
    label.pack(pady=4)

    # Создание кнопок
    Data = tk.Button(menu, text="Database",command=data, width=13, height=2,background="#FF4500", foreground="white",font=('Verdana', 20, 'normal'), cursor="hand2" )
    Data.pack(pady=10)
    Documentation = tk.Button(menu, text="Documentation", command=documentation, width=13, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Documentation.pack(pady=10)
    Exit = tk.Button(menu, text="Exit", command=close, width=13, height=2, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2")
    Exit.pack(pady=10)
    # menu.mainloop()

users = None
# Функция для сохранения пользователей в файл
def save_users():
    with open('users.pickle', 'wb') as file:
        pickle.dump(users, file)

# Загрузка пользователей из файла или создание нового словаря, если файла нет
try:
    with open('users.pickle', 'rb') as file:
        users = pickle.load(file)
except (FileNotFoundError, EOFError):
    users = {"admin": "password"}

def reg():
    def register():
        username = entry_username.get()
        password = entry_password.get()

        # Регистрация нового пользователя
        if username not in users:
            users[username] = password
            save_users()  # Сохраняем пользователей после добавления нового
            label_status.config(text="Пользователь зарегистрирован.")
        else:
            label_status.config(text="Пользователь уже существует.")

    root8 = tk.Tk()
    root8.title("Registration")
    root8.geometry("200x170")
    root8.configure(bg="#0a0a0a")

    label_username = tk.Label(root8, text="", bg="#0a0a0a")
    label_username.pack()

    label_username = tk.Label(root8, text="Username:", bg="#0a0a0a", foreground="white")
    label_username.pack()
    entry_username = tk.Entry(root8)
    entry_username.pack()

    label_password = tk.Label(root8, text="Password:", bg="#0a0a0a", foreground="white")
    label_password.pack()
    entry_password = tk.Entry(root8, show="*")
    entry_password.pack()

    button_register = tk.Button(root8, text="Register", command=register, bg="#FF4500", foreground="white",borderwidth="2")
    button_register.pack(pady=10)

    label_status = tk.Label(root8, text="", bg="#0a0a0a", foreground="white")
    label_status.pack()

    root8.mainloop()

def log():
    def login():
        username = entry_username.get()
        password = entry_password.get()

        # Проверка логина и пароля
        if username in users and users[username] == password:
            main_menu()
            root7.destroy()
            entrance.destroy()
        else:
            label_status.config(text="Invalid username or password")

    root7 = tk.Tk()
    root7.title("Login")
    root7.geometry("200x170")
    root7.configure(bg="#0a0a0a")

    label_username = tk.Label(root7, text="", bg="#0a0a0a")
    label_username.pack()

    label_username = tk.Label(root7, text="Username:", bg="#0a0a0a", foreground="white")
    label_username.pack()
    entry_username = tk.Entry(root7)
    entry_username.pack()

    label_password = tk.Label(root7, text="Password:", bg="#0a0a0a", foreground="white")
    label_password.pack()
    entry_password = tk.Entry(root7, show="*")
    entry_password.pack()

    button_login = tk.Button(root7, text="Enter", command=login, bg="#FF4500", foreground="white", borderwidth="2")
    button_login.pack(pady=10)

    label_status = tk.Label(root7, text="", bg="#0a0a0a", foreground="white")
    label_status.pack()

    root7.mainloop()

def close_add_ed():
    entrance.destroy()
# Начало работы, то, что с самого начал выводится на экран:
# Создание главного окна

entrance = tk.Tk()
entrance.title("Bratishki")
entrance.geometry("300x300")
entrance.resizable(False, False)
entrance.configure(bg="#0a0a0a")
label = tk.Label(entrance, text="   Bratishki Corporation️", font=('Arial', 14, 'normal'), background="#0a0a0a", foreground="white")
label.pack(pady=10)
login1=tk.Button(entrance, text ="Login", width=10, height=1, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2", command=log)
login1.pack(pady=10)
register1=tk.Button(entrance, text ="Registration", width=10, height=1, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2", command=reg)
register1.pack(pady=10)
exit1=tk.Button(entrance, text ="Exit", width=10, height=1, background="#FF4500", foreground="white", font=('Verdana', 20, 'normal'), cursor="hand2", command=close_add_ed)
exit1.pack(pady=10)
entrance.mainloop()
conn.close()
